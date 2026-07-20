"""Generate data/PMO_Master.xlsx with realistic Executive Portfolio data.

Adds executive-grade fields and new sheets so the Enterprise modules work
out-of-the-box. All derived columns use LIVE Excel formulas so the workbook
remains the source of truth and any user-edit recalculates dependents.

Sheets written
--------------
  Projects, Roadmap, Risks, Governance, Financials, Resources, Dependencies,
  Pipeline, CostBenefit                       (existing)
  Benefits, Decisions, Actions, Milestones,
  StageGates, PortfolioMovements, Prioritisation, RAID   (new)

Formulas
--------
  Projects.Remaining Budget          = Approved Funding - Actual Spend
  Projects.Forecast Variance         = Approved Funding - Forecast At Completion
  Projects.Benefits Remaining        = Benefits Forecast - Benefits Realised
  Projects.Benefit Realisation %     = Benefits Realised / Benefits Forecast
  Projects.Total Funding             = CAPEX + OPEX
  Financials.Variance                = Forecast - Actual
  CostBenefit.Total Cost             = CAPEX + OPEX
  CostBenefit.Total Benefit          = Recurring + One-Off
  CostBenefit.Net Benefit            = Total Benefit - Total Cost
  Pipeline.Priority Score            = (Strat*0.3 + Value*0.4 - Risk*0.15 - Effort*0.15)*20
  Governance.Next Gate               = INDEX/MATCH lookup
  Benefits.Benefits Remaining        = Target - Realised
  Benefits.Realisation %             = Realised / Target
  Prioritisation.Score               = Strategic*0.30 + Benefit*0.25 + Risk Reduction*0.20
                                        + Compliance*0.15 - Complexity*0.10  (each /5)
"""
from pathlib import Path
import random
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment

random.seed(42); np.random.seed(42)
OUT = Path(__file__).parent / "data" / "PMO_Master.xlsx"
OUT.parent.mkdir(parents=True, exist_ok=True)

# ─────────────────────────── reference lists ───────────────────────────
PROGRAMS  = ["Digital Transformation","Cyber Security","Cloud Migration","Data & AI","Customer Experience"]
THEMES    = ["Infrastructure","Digital & Applications","Data & Analytics","Cyber Security","Business Transformation"]
SPONSORS  = ["CIO","CFO","COO","CMO","CTO"]
DELIVERY_LEADS = ["John Smith","Jane Doe","Mike Johnson","Sarah Lee","Alex Brown","Priya Patel","Carlos Diaz"]
PMS = DELIVERY_LEADS
BUSINESS_UNITS = ["Retail","Wholesale","Operations","Technology","Finance","HR","Marketing"]
STATUSES  = ["Active","Active","Active","Completed","Pipeline"]
RAGS      = ["Green","Green","Green","Amber","Amber","Red"]
PRIORITY  = ["P1 - Critical","P2 - High","P3 - Medium","P4 - Low"]
INVEST    = ["Run","Grow","Transform"]
FUNDING_TYPES = ["CAPEX","OPEX","Mixed","Unfunded"]
PORTFOLIO_CATEGORIES = ["Business Strategic","IT Strategic","CAPEX","Unfunded"]
FYS       = ["FY24","FY25","FY26","FY27"]
# Multi-FY choices offered on the Projects sheet. A value like "FY27-28"
# means the project's budget is split evenly across FY27 and FY28 in the
# FYAllocation sheet (user can revise the split manually afterwards).
FY_CHOICES = [
    "FY25","FY25-26","FY26","FY26-27","FY27","FY27-28",
    "FY28","FY28-29","FY29","FY29-30","FY30",
]

def _fy_choice_to_list(choice: str) -> list[str]:
    """Expand 'FY27' -> ['FY27']; 'FY27-28' -> ['FY27','FY28']; 'FY28-30' -> ['FY28','FY29','FY30']."""
    s = str(choice or "").strip().upper().replace(" ", "")
    if not s.startswith("FY"):
        return []
    body = s[2:]
    if "-" not in body:
        try: return [f"FY{int(body):02d}"]
        except Exception: return []
    a, b = body.split("-", 1)
    try:
        y0 = int(a); y1 = int(b)
        # allow 2-digit end (FY27-28) or 4-digit
        if y1 < 100 and y0 < 100 and y1 < y0:  # wrap safety
            y1 += 100
        return [f"FY{y:02d}" for y in range(y0, y1 + 1)]
    except Exception:
        return []
DELIVERY_METHODS = ["Waterfall","Waterfall","Agile","Agile","Hybrid"]

# Channel A (<$200K) and Channel B (>$200K) stages
STAGES_A = ["Discovery","Business Case / Full Funding","Design","Build","Testing","Deployment","Handover"]
STAGES_B = ["Discovery","Business Case / Seed Funding","Design","Business Case / Full Funding",
            "Build","Testing","Deployment","Handover"]
STAGES   = STAGES_B  # superset used for lookup
SKILLS   = ["PM","BA","Dev","QA","DevOps","Data","Security","UX","Architect"]

PROJECT_NAMES = [
    "Customer Portal","Cloud Migration","Data Platform","AI Assistant","Network Refresh",
    "Cyber Security Program","CRM Upgrade","Mobile App","ERP Modernisation","Workforce Enablement",
    "IoT Initiative","Analytics Platform","API Gateway","Identity Mgmt","Service Desk",
    "Infrastructure Upg","App Modernisation","RPA Program","Vendor Mgmt","Process Automation",
    "Data Warehouse","DevOps Tools","Workforce Portal","Process Mining","Customer Insights",
    "ERP Modernisation 2","Cyber Program","Mobile App 2","Analytics Platform 2","Service Desk 2",
    "Quality Mgmt","Compliance Hub","Risk Platform","Treasury Sys","Payroll Cloud",
    "HR Self-Service","SAP S/4HANA","Salesforce CRM","M365 Rollout","Zero-Trust Net",
]

# ───────────────────────────── PROJECTS ─────────────────────────────
projects = []
for i, name in enumerate(PROJECT_NAMES, start=1):
    approved = round(random.uniform(0.1, 4.0) * 1_000_000, 0)  # $100K – $4M
    capex    = round(approved * random.uniform(0.5, 0.85), 0)
    opex     = round(approved - capex, 0)
    actual   = round(approved * random.uniform(0.10, 1.05), 0)
    forecast = round(approved * random.uniform(0.85, 1.20), 0)
    ben_forecast = round(approved * random.uniform(1.2, 3.0), 0)
    ben_realised = round(ben_forecast * random.uniform(0.0, 0.6), 0)
    progress = random.randint(5, 100)
    start = pd.Timestamp("2025-01-01") + pd.Timedelta(days=random.randint(0, 365))
    end   = start + pd.Timedelta(days=random.randint(90, 600))
    go_live = end - pd.Timedelta(days=random.randint(0, 60))
    channel  = "Channel A (<$200K)" if approved < 200_000 else "Channel B (>$200K)"
    cat      = random.choice(PORTFOLIO_CATEGORIES)
    funding  = "Unfunded" if cat == "Unfunded" else random.choice(["CAPEX","OPEX","Mixed"])
    projects.append({
        "Project ID":  f"PRJ{i:03d}",
        "Project Name": name,
        "Program":     random.choice(PROGRAMS),
        "Theme":       random.choice(THEMES),
        "Portfolio Category": cat,
        "Business Unit": random.choice(BUSINESS_UNITS),
        "Sponsor":     random.choice(SPONSORS),
        "Delivery Lead": random.choice(DELIVERY_LEADS),
        "PM":          random.choice(PMS),
        "Priority":    random.choice(PRIORITY),
        "Investment Type": random.choice(INVEST),
        "Delivery Method": random.choice(DELIVERY_METHODS),
        "Funding Type": funding,
        "Governance Channel": channel,
        "Financial Year": (lambda sy, ey: f"FY{sy:02d}" if sy == ey else f"FY{sy:02d}-{ey:02d}")(
            int(str(start.year)[-2:]), int(str(end.year)[-2:])),
        "Start Date":  start,
        "End Date":    end,
        "Go Live Date": go_live,
        "Progress %":  progress,
        "Approved Funding": approved,
        "CAPEX":       capex,
        "OPEX":        opex,
        "Total Funding": capex + opex,         # formula later
        "Actual Spend": actual,
        "Forecast At Completion": forecast,
        "Remaining Budget": approved - actual, # formula later
        "Forecast Variance": approved - forecast,  # formula later
        "Contingency": round(approved * 0.10, 0),
        "Benefits Forecast": ben_forecast,
        "Benefits Realised": ben_realised,
        "Benefits Remaining": ben_forecast - ben_realised,    # formula later
        "Benefit Realisation %": round(ben_realised / ben_forecast, 3) if ben_forecast else 0,
        # legacy aliases (kept so existing pages keep working)
        "Budget":      approved,
        "Actual Cost": actual,
        "Forecast":    forecast,
        "Status":      random.choice(STATUSES),
        "RAG":         random.choice(RAGS),
    })
df_projects = pd.DataFrame(projects)

# ───────────────────────────── ROADMAP ─────────────────────────────
roadmap_rows = []
phases = ["Discovery","Design","Build","Test","Deploy"]
for p in projects:
    months = pd.date_range(p["Start Date"], p["End Date"], freq="MS")
    for m in months:
        roadmap_rows.append({"Project ID": p["Project ID"], "Month": m.strftime("%b"),
                             "Year": m.year, "Phase": random.choice(phases),
                             "Status": p["Status"]})
df_roadmap = pd.DataFrame(roadmap_rows)

# ───────────────────────────── RISKS ─────────────────────────────
risks = []
for r in range(1, 61):
    risks.append({
        "Risk ID":     f"R{r:03d}",
        "Project ID":  random.choice(df_projects["Project ID"].tolist()),
        "Description": random.choice([
            "Vendor delay impacting milestone","Budget overrun forecast",
            "Resource shortage","Cyber security vulnerability detected",
            "Scope creep","Integration risk","Regulatory change",
            "Stakeholder alignment risk","Technology obsolescence"]),
        "Probability": random.choice(["Low","Medium","High"]),
        "Impact":      random.choice(["Low","Medium","High","Critical"]),
        "Velocity":    random.choice(["Slow","Medium","Fast"]),
        "Owner":       random.choice(PMS),
        "Mitigation Plan": random.choice([
            "Add buffer","Engage vendor","Rebaseline budget","Patch & monitoring",
            "Increase staffing","Stakeholder workshop"]),
        "Status":      random.choice(["Open","Open","Mitigated","Closed"]),
    })
df_risks = pd.DataFrame(risks)

# ───────────────────────────── RAID (combined) ─────────────────────────────
raid_rows = []
for i in range(1, 81):
    raid_rows.append({
        "RAID ID":    f"RAID{i:03d}",
        "Project ID": random.choice(df_projects["Project ID"].tolist()),
        "Type":       random.choice(["Risk","Issue","Assumption","Dependency"]),
        "Description": random.choice([
            "Cloud provider outage risk","Test data not available",
            "Assumes API contract stable","Depends on procurement signoff",
            "Vendor licence cost rise","Staff turnover","Regulatory deadline"]),
        "Probability": random.choice(["Low","Medium","High"]),
        "Impact":      random.choice(["Low","Medium","High","Critical"]),
        "RAG":         random.choice(["Green","Amber","Red"]),
        "Owner":       random.choice(PMS),
        "Target Resolution Date":
            pd.Timestamp("2026-03-01") + pd.Timedelta(days=random.randint(0, 240)),
        "Mitigation":  random.choice([
            "Active monitoring","Contingency plan","Vendor escalation",
            "Re-baseline","Add buffer"]),
        "Status":      random.choice(["Open","In Progress","Closed"]),
    })
df_raid = pd.DataFrame(raid_rows)

# ─────────── STAGE GATES (unified — replaces old separate Governance sheet) ─
# One row per project per stage. The row whose Stage == current phase carries
# the "current gate" metadata that used to live on the Governance sheet
# (Next Gate, Gate Status, Gate Owner, Gate Outcome, Checklist %).
sg_rows = []
for p in projects:
    stages = STAGES_A if p["Governance Channel"].startswith("Channel A") else STAGES_B
    base = pd.Timestamp(p["Start Date"])
    cur = random.choice(stages)
    cur_idx = stages.index(cur)
    for i, s in enumerate(stages):
        planned = base + pd.Timedelta(days=60*i)
        slip = random.randint(-15, 30)
        is_past = i < cur_idx
        is_current = i == cur_idx
        actual = planned + pd.Timedelta(days=slip) if is_past else pd.NaT
        gate_status = ("Approved" if is_past
                       else random.choice(["Pending Approval","In Progress"]) if is_current
                       else "Not Started")
        sg_rows.append({
            "Project ID":            p["Project ID"],
            "Project Name":          p["Project Name"],
            "Governance Channel":    p["Governance Channel"],
            "Stage":                 s,
            "Next Gate":             stages[min(i+1, len(stages)-1)],
            "Gate Status":           gate_status,
            "Status":                "Complete" if is_past else ("In Progress" if is_current else "Not Started"),
            "Gate Owner":            p["Sponsor"],
            "Planned Gate Date":     planned,
            "Actual Gate Date":      actual,
            "Gate Outcome":          random.choice(["Proceed","Conditional","Rework","Held"]) if is_past else "",
            "Gate Comments":         random.choice(["","Approved with conditions",
                                                    "Awaiting sponsor sign-off","On track"]),
            "Checklist Complete %":  100 if is_past else (random.randint(40,95) if is_current else 0),
            "Days Late":             slip if pd.notna(actual) and slip > 0 else 0,
        })
df_stagegates = pd.DataFrame(sg_rows)

# ───────────────────────────── FINANCIALS (monthly) ─────────────────────────
fin_rows = []
months = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
for p in projects:
    for m in months:
        capex = round(p["CAPEX"]/12 * random.uniform(0.5, 1.5), 0)
        opex  = round(p["OPEX"]/12 * random.uniform(0.5, 1.5), 0)
        actual = round((capex+opex) * random.uniform(0.6, 1.1), 0)
        forecast = round((capex+opex) * random.uniform(0.9, 1.15), 0)
        fin_rows.append({"Project ID": p["Project ID"], "Month": m,
                         "CAPEX": capex, "OPEX": opex,
                         "Actual": actual, "Forecast": forecast,
                         "Variance": forecast - actual,
                         "Planned Value": round((capex+opex) * random.uniform(0.8,1.0),0),
                         "Earned Value":  round((capex+opex) * random.uniform(0.6,1.0),0)})
df_fin = pd.DataFrame(fin_rows)

# ───────────────────────────── RESOURCES (monthly allocation) ─────────────
# One row per (Resource, Project, Month) so allocation trends over time.
res_rows = []
name_to_pid = dict(zip(df_projects["Project Name"], df_projects["Project ID"]))
name_to_start = dict(zip(df_projects["Project Name"], df_projects["Start Date"]))
name_to_end   = dict(zip(df_projects["Project Name"], df_projects["End Date"]))
for r in range(1, 31):
    name = f"Resource {r:02d}"
    skill = random.choice(SKILLS)
    role = random.choice(["Project Manager","Developer","Tester","Analyst","Architect"])
    for _ in range(random.randint(1, 3)):
        proj = random.choice(df_projects["Project Name"].tolist())
        base_alloc = random.choice([25, 50, 75, 100])
        # Monthly rows across project window (fallback: 12 months from 2025-01)
        p_start = pd.Timestamp(name_to_start.get(proj, pd.Timestamp("2025-01-01")))
        p_end   = pd.Timestamp(name_to_end.get(proj,   pd.Timestamp("2025-12-01")))
        months = pd.date_range(p_start.replace(day=1), p_end, freq="MS")
        if len(months) == 0:
            months = pd.date_range("2025-01-01", periods=12, freq="MS")
        for m in months:
            # gentle variation month-to-month, capped 0..100
            alloc = max(0, min(100, base_alloc + random.choice([-25, -10, 0, 0, 10, 15])))
            res_rows.append({
                "Resource":     name,
                "Skill":        skill,
                "Role":         role,
                "Project ID":   name_to_pid.get(proj, ""),
                "Project":      proj,
                "Month":        m,
                "Allocation %": alloc,
                "Capacity %":   100,
            })
df_res = pd.DataFrame(res_rows)


# ───────────────────────────── DEPENDENCIES ─────────────────────────────
dep_rows = []
ids = df_projects["Project ID"].tolist()
for _ in range(40):
    a, b = random.sample(ids, 2)
    dep_rows.append({"From Project": a, "To Project": b,
                     "Dependency Type": random.choice(["Finish-Start","Start-Start","Blocks","Shared Resource"]),
                     "Status": random.choice(["Healthy","At Risk","Blocked"]),
                     "Impact": random.choice(["Low","Medium","High"])})
df_dep = pd.DataFrame(dep_rows)

# ───────────────────────────── PIPELINE ─────────────────────────────
pipeline_rows = []
ideas = ["Process Automation v2","Customer 360","ML Ops Platform","Edge Analytics","Green IT",
         "ESG Data Hub","Citizen Dev Studio","Quantum Pilot","SaaS Rationalisation","Zero Trust v2"]
for i, idea in enumerate(ideas, start=1):
    strat = random.randint(1,5); value = random.randint(1,5); risk = random.randint(1,5); effort = random.randint(1,5)
    score = round((strat*0.3 + value*0.4 - risk*0.15 - effort*0.15) * 20, 1)
    pipeline_rows.append({"Idea ID": f"IDEA{i:03d}", "Name": idea,
                          "Submitter": random.choice(PMS),
                          "Strategic Fit (1-5)": strat, "Value (1-5)": value,
                          "Risk (1-5)": risk, "Effort (1-5)": effort,
                          "Priority Score": score,
                          "Decision": random.choice(["New","Under Review","Approved","Rejected","Parked"]),
                          "Est. Budget ($K)": random.randint(50, 2000)})
df_pipe = pd.DataFrame(pipeline_rows).sort_values("Priority Score", ascending=False).reset_index(drop=True)

# ───────────────────────────── COST vs BENEFIT ─────────────────────────────
# Year window is anchored to each project's Start Date. Recurring benefits
# and OPEX run for 5 years starting with the project's start year, so a
# project that starts in 2026 publishes CostBenefit rows for 2026-2030 —
# never earlier. The Year cell is also written as a formula tied to
# Projects.Start Date so editing the start date re-flows the years.
BENEFIT_CATEGORIES = ["Cost Savings","Revenue Uplift","Productivity","Risk Avoidance","Compliance"]
CB_HORIZON_YEARS = 5
cb_rows = []
for p in projects:
    btype = random.choice(["Recurring","Recurring","One-Off"])
    bcat  = random.choice(BENEFIT_CATEGORIES)
    total_capex = p["Approved Funding"]
    yearly_opex = round(total_capex * random.uniform(0.05, 0.15), 0)
    base_benefit = round(total_capex * random.uniform(0.15, 0.65), 0)
    one_off_year = random.choice([2,3])
    start_year = pd.Timestamp(p["Start Date"]).year
    for idx in range(1, CB_HORIZON_YEARS + 1):
        yr = start_year + (idx - 1)
        if idx == 1:   capex = round(total_capex * random.uniform(0.55, 0.75), 0)
        elif idx == 2: capex = round(total_capex * random.uniform(0.20, 0.40), 0)
        else:          capex = round(total_capex * random.uniform(0.00, 0.05), 0)
        opex = 0 if idx == 1 else round(yearly_opex * random.uniform(0.8, 1.2), 0)
        ben_rec = ben_one = 0
        if btype == "Recurring" and idx >= 2:
            ben_rec = round(base_benefit * random.uniform(0.8, 1.2), 0)
        if btype == "One-Off" and idx == one_off_year:
            ben_one = round(base_benefit * random.uniform(1.5, 2.5), 0)
        cb_rows.append({
            "Project ID":        p["Project ID"],
            "Project Name":      p["Project Name"],
            "Program":           p["Program"],
            "Year Offset":       idx - 1,
            "Year":              yr,
            "CAPEX":             capex,
            "OPEX":              opex,
            "Total Cost":        capex + opex,
            "Benefit Recurring": ben_rec,
            "Benefit One-Off":   ben_one,
            "Total Benefit":     ben_rec + ben_one,
            "Net Benefit":       (ben_rec + ben_one) - (capex + opex),
            "Benefit Type":      btype,
            "Benefit Category":  bcat,
            "Confidence %":      random.choice([60,70,80,90]),
        })
df_costbenefit = pd.DataFrame(cb_rows)

# ───────────────────────────── BENEFITS REGISTER ─────────────────────────────
ben_rows = []
ben_id = 1
for p in projects:
    for _ in range(random.randint(1, 3)):
        target = round(random.uniform(0.1, 2.0) * 1_000_000, 0)
        realised = round(target * random.uniform(0, 0.7), 0)
        ben_rows.append({
            "Benefit ID":   f"BEN{ben_id:03d}",
            "Project ID":   p["Project ID"],
            "Project Name": p["Project Name"],
            "Description":  random.choice([
                "Headcount cost avoidance","Revenue growth from new channel",
                "Process cycle-time reduction","Risk loss avoidance",
                "Regulatory compliance achievement","Customer NPS uplift"]),
            "Category":     random.choice(["Revenue Increase","Cost Reduction",
                                          "Productivity Improvement","Customer Experience",
                                          "Risk Reduction","Compliance"]),
            "Target Value": target,
            "Realised Value": realised,
            "Benefits Remaining": target - realised,        # formula later
            "Realisation %": round(realised/target, 3) if target else 0,  # formula later
            "Owner":        random.choice(PMS),
            "Start Date":   pd.Timestamp("2025-06-01") + pd.Timedelta(days=random.randint(0, 180)),
            "End Date":     pd.Timestamp("2027-06-01") + pd.Timedelta(days=random.randint(0, 365)),
            "Status":       random.choice(["Planned","In Progress","Realised","At Risk"]),
        })
        ben_id += 1
df_benefits = pd.DataFrame(ben_rows)

# ───────────────────────────── DECISIONS ─────────────────────────────
dec_rows = []
DEC_TYPES = ["Project","Program","Portfolio","Funding","SteerCo","Architecture","Risk","General"]
for i in range(1, 41):
    due = pd.Timestamp("2026-02-01") + pd.Timedelta(days=random.randint(-30, 200))
    dec_rows.append({
        "Decision ID":    f"DEC{i:03d}",
        "Type":           random.choice(DEC_TYPES),
        "Project ID":     random.choice(df_projects["Project ID"].tolist() + [""]*4),
        "Program":        random.choice(PROGRAMS + [""]*3),
        "Description":    random.choice([
            "Approve full project funding","Choose vendor for cloud migration",
            "Cut scope of MVP","Re-baseline schedule","Approve architecture pattern",
            "Accept residual risk","Defer benefits realisation date"]),
        "Required By":    random.choice(SPONSORS),
        "Owner":          random.choice(PMS),
        "Due Date":       due,
        "Priority":       random.choice(["High","Medium","Low"]),
        "Status":         random.choice(["Open","In Review","Approved","Rejected","Closed"]),
        "Outcome":        "",
        "Approval Date":  pd.NaT,
    })
df_dec = pd.DataFrame(dec_rows)

# ───────────────────────────── ACTIONS ─────────────────────────────
act_rows = []
for i in range(1, 61):
    due = pd.Timestamp("2026-01-15") + pd.Timedelta(days=random.randint(-30, 180))
    status = random.choice(["Open","In Progress","Complete"])
    if due < pd.Timestamp.today() and status != "Complete":
        status = "Overdue"
    act_rows.append({
        "Action ID":   f"ACT{i:03d}",
        "Project ID":  random.choice(df_projects["Project ID"].tolist()),
        "Description": random.choice([
            "Confirm vendor SOW","Update risk register","Publish SteerCo pack",
            "Resolve blocker on environment","Schedule architecture review",
            "Submit funding paper","Close out stage gate evidence"]),
        "Owner":       random.choice(PMS),
        "Due Date":    due,
        "Priority":    random.choice(["High","Medium","Low"]),
        "Status":      status,
        "Comments":    "",
    })
df_act = pd.DataFrame(act_rows)

# ───────────────────────────── MILESTONES ─────────────────────────────
# Derived from StageGates: one milestone row per stage gate. Users can
# append manual rows (set Source = "Manual") without breaking auto-fill.
ms_rows = []
for row in sg_rows:
    ms_rows.append({
        "Project ID":    row["Project ID"],
        "Milestone":     row["Stage"],
        "Planned Date":  row["Planned Gate Date"],
        "Forecast Date": row["Actual Gate Date"] if pd.notna(row["Actual Gate Date"]) else row["Planned Gate Date"],
        "Actual Date":   row["Actual Gate Date"],
        "Status":        row["Status"],
        "Owner":         row["Gate Owner"],
        "Source":        "StageGate",
    })
df_ms = pd.DataFrame(ms_rows)


# ───────────────────────────── PORTFOLIO MOVEMENTS (audit trail) ─────────
mv_rows = []
for i in range(1, 11):
    p = random.choice(projects)
    old = random.choice(PORTFOLIO_CATEGORIES)
    new = random.choice([c for c in PORTFOLIO_CATEGORIES if c != old])
    mv_rows.append({
        "Movement ID": f"MV{i:03d}",
        "Project ID":  p["Project ID"],
        "Project Name": p["Project Name"],
        "From Category": old,
        "To Category":   new,
        "Moved By":      random.choice(["admin","governance","sponsor"]),
        "Moved On":      pd.Timestamp("2026-01-01") + pd.Timedelta(days=random.randint(-180, 60)),
        "Reason":        random.choice([
            "Funding approved — move to CAPEX",
            "Reclassified following strategy review",
            "De-funded; moved to unfunded backlog",
            "Promoted from unfunded to IT strategic",
        ]),
    })
df_mv = pd.DataFrame(mv_rows)

# ───────────────────────────── PRIORITISATION SCORING ─────────────────────
pr_rows = []
for p in projects:
    strat = random.randint(1,5); benefit = random.randint(1,5)
    risk_red = random.randint(1,5); compliance = random.randint(1,5)
    complex_ = random.randint(1,5)
    score = round((strat*0.30 + benefit*0.25 + risk_red*0.20 + compliance*0.15
                   - complex_*0.10) * 20, 1)
    pr_rows.append({
        "Project ID":   p["Project ID"],
        "Project Name": p["Project Name"],
        "Strategic Alignment (1-5)":   strat,
        "Benefit Value (1-5)":         benefit,
        "Risk Reduction (1-5)":        risk_red,
        "Compliance (1-5)":            compliance,
        "Complexity (1-5)":            complex_,
        "Score":                       score,
    })
df_prior = pd.DataFrame(pr_rows).sort_values("Score", ascending=False).reset_index(drop=True)

# ───────────────────────────── CONFIG (single source of truth) ────────────
# Config sheet: every dropdown list the app uses, fully user-editable.
CONFIG_LISTS = {
    "Status":              STATUSES + ["On Hold","Cancelled"],
    "RAG":                 ["Green","Amber","Red"],
    "Priority":            PRIORITY,
    "Investment Type":     INVEST,
    "Delivery Method":     ["Waterfall","Agile","Hybrid"],
    "Funding Type":        FUNDING_TYPES,
    "Portfolio Category":  PORTFOLIO_CATEGORIES,
    "Governance Channel":  ["Channel A (<$200K)","Channel B (>$200K)"],
    "Program":             PROGRAMS,
    "Theme":               THEMES,
    "Business Unit":       BUSINESS_UNITS,
    "Sponsor":             SPONSORS,
    "Financial Year":      FY_CHOICES,
    "Stage (Channel A)":   STAGES_A,
    "Stage (Channel B)":   STAGES_B,
    "Gate Status":         ["Not Started","In Progress","Pending Approval",
                            "Approved","Rejected","Complete","On Hold"],
    "Gate Outcome":        ["","Proceed","Conditional","Rework","Held"],
    "Decision Status":     ["Open","In Review","Approved","Rejected","Closed"],
    "Decision Type":       DEC_TYPES,
    "Action Status":       ["Open","In Progress","Complete","Overdue"],
    "Risk Probability":    ["Low","Medium","High"],
    "Risk Impact":         ["Low","Medium","High","Critical"],
    "Risk Status":         ["Open","Mitigated","Closed"],
    "RAID Type":           ["Risk","Issue","Assumption","Dependency"],
    "Benefit Status":      ["Planned","In Progress","Realised","At Risk"],
    "Benefit Category":    BENEFIT_CATEGORIES,
    "Benefit Type":        ["Recurring","One-Off"],
    "Dependency Type":     ["Finish-Start","Start-Start","Blocks","Shared Resource"],
    "Dependency Status":   ["Healthy","At Risk","Blocked"],
    "Skill":               SKILLS,
    "Resource Role":       ["Project Manager","Developer","Tester","Analyst","Architect"],
    "Milestone":           STAGES,
    "Pipeline Decision":   ["New","Under Review","Approved","Rejected","Parked"],
}
cfg_rows = []
for cat, vals in CONFIG_LISTS.items():
    for i, v in enumerate(vals, start=1):
        cfg_rows.append({"Category": cat, "Value": v,
                         "Display Order": i, "Active": True})
df_config = pd.DataFrame(cfg_rows)

# Rules sheet: every threshold the app uses. Edit here → app re-reads.
RULES = [
    ("GOV_CHANNEL_THRESHOLD",     200000,  "currency", "Approved $ at/below which a project is Channel A"),
    ("SCHEDULE_AMBER_VARIANCE",   -0.05,   "ratio",    "Schedule variance worse than this → Amber"),
    ("SCHEDULE_RED_VARIANCE",     -0.10,   "ratio",    "Schedule variance worse than this → Red"),
    ("FINANCIAL_AMBER_VARIANCE",   0.05,   "ratio",    "(Forecast-Approved)/Approved above → Amber"),
    ("FINANCIAL_RED_VARIANCE",     0.10,   "ratio",    "(Forecast-Approved)/Approved above → Red"),
    ("BENEFIT_GREEN_THRESHOLD",    0.70,   "ratio",    "Benefit realisation ≥ → Green"),
    ("BENEFIT_AMBER_THRESHOLD",    0.30,   "ratio",    "Benefit realisation ≥ → Amber (else Red)"),
    ("DELIVERY_AMBER_DAYS_LATE",   1,      "int",      "Stage gate days late ≥ → Amber"),
    ("DELIVERY_RED_DAYS_LATE",     15,     "int",      "Stage gate days late ≥ → Red"),
    ("UPCOMING_GATE_WINDOW_DAYS",  30,     "int",      "Window for 'upcoming stage gates' KPI"),
    ("PRIORITY_SCORE_WEIGHT_STRAT",  0.30, "ratio",    "Prioritisation: strategic weight"),
    ("PRIORITY_SCORE_WEIGHT_BENEFIT",0.25, "ratio",    "Prioritisation: benefit weight"),
    ("PRIORITY_SCORE_WEIGHT_RISK",   0.20, "ratio",    "Prioritisation: risk-reduction weight"),
    ("PRIORITY_SCORE_WEIGHT_COMPL",  0.15, "ratio",    "Prioritisation: compliance weight"),
    ("PRIORITY_SCORE_WEIGHT_CPLX",   0.10, "ratio",    "Prioritisation: complexity penalty"),
    ("CURRENCY_SYMBOL",            "$",    "text",     "Symbol shown in KPI cards"),
]
df_rules = pd.DataFrame(RULES, columns=["Key","Value","Type","Description"])

# ───────────────────────────── SPRINTS (agile projects) ─────────────────
# One row per sprint per agile/hybrid project. Two-week sprints, 6 sprints each.
sprint_rows = []
today = pd.Timestamp.today().normalize()
sid = 1
for p in projects:
    if p["Delivery Method"] not in ("Agile", "Hybrid"):
        continue
    team_velocity = random.randint(25, 55)  # baseline points/sprint
    base = pd.Timestamp(p["Start Date"])
    for n in range(1, 9):  # 8 sprints
        s_start = base + pd.Timedelta(days=(n - 1) * 14)
        s_end   = s_start + pd.Timedelta(days=13)
        committed = team_velocity + random.randint(-8, 10)
        # Sprint status vs today
        if s_end < today:
            status = "Complete"
            completed = max(0, committed + random.randint(-12, 5))
        elif s_start <= today <= s_end:
            status = "Active"
            frac = (today - s_start).days / 14.0
            completed = int(committed * frac * random.uniform(0.6, 1.05))
        else:
            status = "Planned"
            completed = 0
        stories_c = max(1, committed // 5)
        stories_done = max(0, completed // 5)
        sprint_rows.append({
            "Sprint ID":       f"SP{sid:04d}",
            "Project ID":      p["Project ID"],
            "Project Name":    p["Project Name"],
            "Sprint #":        n,
            "Sprint Name":     f"{p['Project ID']} Sprint {n}",
            "Start Date":      s_start,
            "End Date":        s_end,
            "Points Committed": committed,
            "Points Completed": completed,
            "Stories Committed": stories_c,
            "Stories Completed": stories_done,
            "Status":          status,
            "Team":            p.get("Delivery Lead", ""),
        })
        sid += 1
df_sprints = pd.DataFrame(sprint_rows)

# ───────────────────────────── RELEASES ─────────────────────────────
release_rows = []
rid = 1
REL_TYPES = ["Major", "Minor", "Patch", "Hotfix"]
REL_ENVS  = ["Production", "UAT", "Staging"]
for p in projects:
    n_rel = random.randint(1, 4)
    span = (pd.Timestamp(p["End Date"]) - pd.Timestamp(p["Start Date"])).days or 1
    for k in range(1, n_rel + 1):
        planned = pd.Timestamp(p["Start Date"]) + pd.Timedelta(days=int(span * k / (n_rel + 1)))
        if planned < today:
            status = "Delivered"
            actual = planned + pd.Timedelta(days=random.randint(-5, 10))
        elif planned < today + pd.Timedelta(days=30):
            status = "In Progress"; actual = pd.NaT
        else:
            status = "Planned"; actual = pd.NaT
        release_rows.append({
            "Release ID":   f"REL{rid:04d}",
            "Project ID":   p["Project ID"],
            "Project Name": p["Project Name"],
            "Release Name": f"{p['Project Name']} R{k}",
            "Version":      f"{k}.0.0",
            "Type":         random.choice(REL_TYPES),
            "Planned Date": planned,
            "Actual Date":  actual,
            "Status":       status,
            "Owner":        p.get("Delivery Lead", ""),
            "Environment":  random.choice(REL_ENVS),
            "Notes":        "",
        })
        rid += 1
df_releases = pd.DataFrame(release_rows)

# ───────────────────────────── FY ALLOCATION ─────────────────────────────
def _fy_label(ts, fy_start_month=1):
    ts = pd.Timestamp(ts)
    y = ts.year if ts.month >= fy_start_month else ts.year - 1
    return f"FY{str(y+1)[-2:]}" if fy_start_month != 1 else f"FY{str(y)[-2:]}"

fy_rows = []
for p in projects:
    # Prefer the explicit multi-FY dropdown value on the project. Fall back
    # to spanning the Start/End dates so blank cells still get an allocation.
    fys_span = _fy_choice_to_list(p.get("Financial Year", ""))
    if not fys_span:
        start = pd.Timestamp(p["Start Date"]); end = pd.Timestamp(p["End Date"])
        fys_span = sorted({_fy_label(d) for d in pd.date_range(start, end, freq="MS")})
    if not fys_span:
        fys_span = [_fy_label(pd.Timestamp(p["Start Date"]))]
    approved = float(p["Approved Funding"]); forecast = float(p["Forecast At Completion"])
    n = len(fys_span)
    for fy in fys_span:
        share = 1.0 / n
        fy_rows.append({
            "Project ID":      p["Project ID"],
            "Project Name":    p["Project Name"],
            "FY":              fy,
            "Budget %":        round(share * 100, 1),
            "Forecast %":      round(share * 100, 1),
            "Budget Amount":   round(approved * share, 0),
            "Forecast Amount": round(forecast * share, 0),
            "Notes":           "Auto-split from Projects.Financial Year — edit to revise",
        })
df_fyalloc = pd.DataFrame(fy_rows)

# ───────────────────────────── PROGRAMS ─────────────────────────────
prog_rows = []
for prog in PROGRAMS:
    sub = df_projects[df_projects["Program"] == prog]
    if sub.empty: continue
    prog_rows.append({
        "Program":  prog,
        "Owner":    random.choice(DELIVERY_LEADS),
        "Sponsor":  random.choice(SPONSORS),
        "Budget":   float(sub["Approved Funding"].sum()),
        "Forecast": float(sub["Forecast At Completion"].sum()),
        "Start FY": _fy_label(sub["Start Date"].min()),
        "End FY":   _fy_label(sub["End Date"].max()),
        "Status":   "Active",
        "Notes":    "",
    })
df_programs = pd.DataFrame(prog_rows)

# ───────────────────────────── PHASE FINANCIALS ─────────────────────────────
phase_rows = []
for p in projects:
    stages = STAGES_A if "Channel A" in p["Governance Channel"] else STAGES_B
    span_days = max(1, (pd.Timestamp(p["End Date"]) - pd.Timestamp(p["Start Date"])).days)
    per = span_days // len(stages)
    approved = float(p["Approved Funding"])
    for i, stg_name in enumerate(stages):
        ps = pd.Timestamp(p["Start Date"]) + pd.Timedelta(days=i * per)
        pe = ps + pd.Timedelta(days=per)
        share = 1.0 / len(stages)
        done = pe < today
        phase_rows.append({
            "Project ID":         p["Project ID"],
            "Stage":              stg_name,
            "Planned Start":      ps,
            "Planned End":        pe,
            "Actual Start":       ps if ps < today else pd.NaT,
            "Actual End":         pe if done else pd.NaT,
            "Phase Budget":       round(approved * share, 0),
            "Phase Forecast":     round(approved * share * random.uniform(0.9, 1.15), 0),
            "Phase Actual Spend": round(approved * share * random.uniform(0.0, 1.05), 0) if ps < today else 0,
            "Status":             "Complete" if done else ("In Progress" if ps < today else "Planned"),
            "Notes":              "",
        })
df_phasefin = pd.DataFrame(phase_rows)

# ───────────────────────────── WRITE WORKBOOK ─────────────────────────────
with pd.ExcelWriter(OUT, engine="openpyxl") as xw:
    df_projects.to_excel(xw,     sheet_name="Projects",            index=False)
    df_roadmap.to_excel(xw,      sheet_name="Roadmap",             index=False)
    df_risks.to_excel(xw,        sheet_name="Risks",               index=False)
    df_raid.to_excel(xw,         sheet_name="RAID",                index=False)
    df_stagegates.to_excel(xw,   sheet_name="StageGates",          index=False)
    df_fin.to_excel(xw,          sheet_name="Financials",          index=False)
    df_res.to_excel(xw,          sheet_name="Resources",           index=False)
    df_dep.to_excel(xw,          sheet_name="Dependencies",        index=False)
    df_pipe.to_excel(xw,         sheet_name="Pipeline",            index=False)
    df_costbenefit.to_excel(xw,  sheet_name="CostBenefit",         index=False)
    df_benefits.to_excel(xw,     sheet_name="Benefits",            index=False)
    df_dec.to_excel(xw,          sheet_name="Decisions",           index=False)
    df_act.to_excel(xw,          sheet_name="Actions",             index=False)
    df_ms.to_excel(xw,           sheet_name="Milestones",          index=False)
    df_mv.to_excel(xw,           sheet_name="PortfolioMovements",  index=False)
    df_prior.to_excel(xw,        sheet_name="Prioritisation",      index=False)
    df_config.to_excel(xw,       sheet_name="Config",              index=False)
    df_rules.to_excel(xw,        sheet_name="ConfigRules",         index=False)
    df_sprints.to_excel(xw,      sheet_name="Sprints",             index=False)
    df_releases.to_excel(xw,     sheet_name="Releases",            index=False)
    df_fyalloc.to_excel(xw,      sheet_name="FYAllocation",        index=False)
    df_programs.to_excel(xw,     sheet_name="Programs",            index=False)
    df_phasefin.to_excel(xw,     sheet_name="PhaseFinancials",     index=False)

    # Project Brief (Sections 1 & 2) — one row per project, populated via UI
    brief_cols = [
        "Project ID", "Portfolio / Workstream", "Sponsor", "Business Owner",
        "Business Solution Manager", "Strategic Alignment",
        "Background and Context", "Opportunity / Problem Statement",
        "Objective", "In Scope", "Out of Scope",
        "Assumptions & Constraints", "Key Metrics / Success Measures",
        "Approval Type", "Funding Ask", "Funding Source", "Resource Ask",
        "Estimate Commentary", "P&L Benefits Commentary",
        "Delivery Milestones", "Project Risks", "Dependencies",
    ]
    df_brief = pd.DataFrame(
        [{**{c: "" for c in brief_cols}, "Project ID": pid}
         for pid in df_projects["Project ID"].tolist()],
        columns=brief_cols,
    )
    df_brief.to_excel(xw, sheet_name="ProjectBrief", index=False)

    # Document links attached to a project (Charter, BRD, Test Plan, etc.)
    df_links = pd.DataFrame(columns=["Project ID", "Title", "Type", "URL", "Added By", "Added On"])
    df_links.to_excel(xw, sheet_name="ProjectLinks", index=False)

# ───────────────────────────── LIVE FORMULAS ─────────────────────────────
wb = load_workbook(OUT)

def col_letters(ws, header_name):
    for cell in ws[1]:
        if cell.value == header_name:
            return get_column_letter(cell.column)
    return None

def set_formula(ws, target, template):
    col = col_letters(ws, target)
    if not col:
        return
    for r in range(2, ws.max_row + 1):
        ws[f"{col}{r}"] = template.format(r=r)

# Projects
ws = wb["Projects"]
appr   = col_letters(ws, "Approved Funding")
actual = col_letters(ws, "Actual Spend")
fc     = col_letters(ws, "Forecast At Completion")
capx   = col_letters(ws, "CAPEX"); opx = col_letters(ws, "OPEX")
bf = col_letters(ws, "Benefits Forecast"); br = col_letters(ws, "Benefits Realised")
set_formula(ws, "Remaining Budget",     f"={appr}{{r}}-{actual}{{r}}")
set_formula(ws, "Forecast Variance",    f"={appr}{{r}}-{fc}{{r}}")
set_formula(ws, "Total Funding",        f"={capx}{{r}}+{opx}{{r}}")
set_formula(ws, "Benefits Remaining",   f"={bf}{{r}}-{br}{{r}}")
set_formula(ws, "Benefit Realisation %",f"=IFERROR({br}{{r}}/{bf}{{r}},0)")

# Financials
ws = wb["Financials"]
f_col = col_letters(ws, "Forecast"); a_col = col_letters(ws, "Actual")
set_formula(ws, "Variance", f"={f_col}{{r}}-{a_col}{{r}}")

# CostBenefit
ws = wb["CostBenefit"]
cap = col_letters(ws, "CAPEX"); opx = col_letters(ws, "OPEX")
br_ = col_letters(ws, "Benefit Recurring"); bo = col_letters(ws, "Benefit One-Off")
tc = col_letters(ws, "Total Cost"); tb = col_letters(ws, "Total Benefit")
set_formula(ws, "Total Cost",    f"={cap}{{r}}+{opx}{{r}}")
set_formula(ws, "Total Benefit", f"={br_}{{r}}+{bo}{{r}}")
set_formula(ws, "Net Benefit",   f"={tb}{{r}}-{tc}{{r}}")

# Year = YEAR(Projects.Start Date) + Year Offset — so editing a project's
# Start Date automatically shifts its 5-year CostBenefit window.
_cb_pid = col_letters(ws, "Project ID")
_cb_yr  = col_letters(ws, "Year")
_cb_off = col_letters(ws, "Year Offset")
_p_start_idx = None
for _c in wb["Projects"][1]:
    if _c.value == "Start Date":
        _p_start_idx = _c.column
        break
if _cb_pid and _cb_yr and _cb_off and _p_start_idx:
    _proj_max = get_column_letter(wb["Projects"].max_column)
    for _r in range(2, ws.max_row + 1):
        ws[f"{_cb_yr}{_r}"] = (
            f'=IFERROR(YEAR(VLOOKUP({_cb_pid}{_r},Projects!$A:${_proj_max},'
            f'{_p_start_idx},FALSE))+{_cb_off}{_r},{_cb_off}{_r})'
        )
    # Hide the offset helper column
    ws.column_dimensions[_cb_off].hidden = True

# Pipeline
ws = wb["Pipeline"]
s  = col_letters(ws, "Strategic Fit (1-5)"); v = col_letters(ws, "Value (1-5)")
rk = col_letters(ws, "Risk (1-5)");          ef = col_letters(ws, "Effort (1-5)")
set_formula(ws, "Priority Score", f"=({s}{{r}}*0.3+{v}{{r}}*0.4-{rk}{{r}}*0.15-{ef}{{r}}*0.15)*20")

# Benefits
ws = wb["Benefits"]
tv = col_letters(ws, "Target Value"); rv = col_letters(ws, "Realised Value")
set_formula(ws, "Benefits Remaining", f"={tv}{{r}}-{rv}{{r}}")
set_formula(ws, "Realisation %",      f"=IFERROR({rv}{{r}}/{tv}{{r}},0)")

# Prioritisation
ws = wb["Prioritisation"]
sa = col_letters(ws, "Strategic Alignment (1-5)")
bv = col_letters(ws, "Benefit Value (1-5)")
rr = col_letters(ws, "Risk Reduction (1-5)")
cp = col_letters(ws, "Compliance (1-5)")
cx = col_letters(ws, "Complexity (1-5)")
set_formula(ws, "Score", f"=({sa}{{r}}*0.30+{bv}{{r}}*0.25+{rr}{{r}}*0.20+{cp}{{r}}*0.15-{cx}{{r}}*0.10)*20")

# ── Auto-fill descriptive fields FROM Projects via VLOOKUP ───────────────
# Editing a value in Projects (Project Name, Program, Sponsor, Portfolio
# Category, Business Unit, Priority, Status, RAG) automatically updates every
# child sheet that carries the same column. Users only maintain Projects.
_proj_ws = wb["Projects"]
_proj_max_col = get_column_letter(_proj_ws.max_column)
_proj_pid_col = col_letters(_proj_ws, "Project ID") or "A"
# Map header -> 1-based column index within Projects
_proj_header_idx = {c.value: c.column for c in _proj_ws[1] if c.value}

# Project-level attributes that flow down to every child sheet. Excludes
# fields that legitimately mean different things per child row (Status/RAG on
# Risks, Actions, Decisions etc.). Any child sheet that has both a "Project
# ID" column and a column named the same as one of these will get a VLOOKUP
# formula so it stays in sync with Projects automatically.
_PROJECT_ATTRS = [
    "Project Name", "Program", "Sponsor", "Portfolio Category",
    "Business Unit", "Funding Type", "Governance Channel", "Priority",
    "Delivery Method", "Delivery Lead", "PM", "Investment Type",
    "Financial Year", "Start Date", "End Date", "Go Live Date",
    "Approved Funding", "CAPEX", "OPEX", "Total Funding",
]
# Sheets excluded from bulk autofill (they own their own attribute columns).
_AUTOFILL_SKIP = {"Projects", "Pipeline", "Programs", "PortfolioMovements",
                  "Dependencies", "Resources", "_Stages"}
# Per-sheet blocklist: columns that MUST NOT be overwritten even though the
# name collides with a Projects attribute (e.g. Milestones.Status is the
# milestone's own status, not the project's).
_AUTOFILL_BLOCK = {
    "Milestones":  {"Status"},
    "Actions":     {"Status", "Priority"},
    "Decisions":   {"Status"},
    "Risks":       {"Status", "Priority"},
    "RAID":        {"Status", "Priority"},
    "StageGates":  {"Status"},
    "Sprints":     {"Status"},
    # CostBenefit/Financials carry their own per-year / per-month CAPEX & OPEX
    # values — do NOT overwrite them with Projects-level totals via VLOOKUP.
    "CostBenefit": {"CAPEX", "OPEX", "Approved Funding", "Total Funding"},
    "Financials":  {"CAPEX", "OPEX", "Approved Funding", "Total Funding"},
}
_AUTOFILL_TARGETS = []
for _sname in wb.sheetnames:
    if _sname in _AUTOFILL_SKIP:
        continue
    _ws_check = wb[_sname]
    if not col_letters(_ws_check, "Project ID"):
        continue
    _block = _AUTOFILL_BLOCK.get(_sname, set())
    _cols = [c for c in _PROJECT_ATTRS
             if col_letters(_ws_check, c) and c not in _block]
    if _cols:
        _AUTOFILL_TARGETS.append((_sname, _cols))
for _sname, _cols in _AUTOFILL_TARGETS:
    if _sname not in wb.sheetnames:
        continue
    _ws = wb[_sname]
    _pid = col_letters(_ws, "Project ID")
    if not _pid or _ws.max_row < 2:
        continue
    for _col_name in _cols:
        _tgt = col_letters(_ws, _col_name)
        _src_idx = _proj_header_idx.get(_col_name)
        if not _tgt or not _src_idx:
            continue
        for _r in range(2, _ws.max_row + 1):
            _ws[f"{_tgt}{_r}"] = (
                f'=IFERROR(VLOOKUP({_pid}{_r},Projects!$A:${_proj_max_col},'
                f'{_src_idx},FALSE),"")'
            )


# ── Milestones ⟵ StageGates (composite key: Project ID + Stage) ───────────
# Add hidden helper column on both sheets so Milestones Planned/Actual dates,
# status and owner track the source stage gate. Manual milestone rows
# (Source column left as "Manual" and no matching stage) fall back to any
# value the user typed — IFERROR keeps the cell empty when no match.
if "StageGates" in wb.sheetnames and "Milestones" in wb.sheetnames:
    _sg = wb["StageGates"]
    _ms = wb["Milestones"]
    _sg_pid = col_letters(_sg, "Project ID")
    _sg_stage = col_letters(_sg, "Stage")
    _ms_pid = col_letters(_ms, "Project ID")
    _ms_mile = col_letters(_ms, "Milestone")
    if _sg_pid and _sg_stage and _ms_pid and _ms_mile:
        # Helper key columns at first empty column on each sheet
        _sg_key = get_column_letter(_sg.max_column + 1)
        _sg.cell(row=1, column=_sg.max_column + 1, value="_Key")
        for _r in range(2, _sg.max_row + 1):
            _sg[f"{_sg_key}{_r}"] = f"={_sg_pid}{_r}&\"|\"&{_sg_stage}{_r}"
        _ms_key = get_column_letter(_ms.max_column + 1)
        _ms.cell(row=1, column=_ms.max_column + 1, value="_Key")
        for _r in range(2, _ms.max_row + 1):
            _ms[f"{_ms_key}{_r}"] = f"={_ms_pid}{_r}&\"|\"&{_ms_mile}{_r}"

        # Use INDEX/MATCH so key column position doesn't matter
        _sg_headers = {c.value: c.column for c in _sg[1] if c.value}
        _key_col_letter = get_column_letter(_sg_headers["_Key"])

        # Milestone target col -> StageGates source header
        _MS_MAP = {
            "Planned Date":  "Planned Gate Date",
            "Actual Date":   "Actual Gate Date",
            "Forecast Date": "Actual Gate Date",
            "Status":        "Status",
            "Owner":         "Gate Owner",
        }
        for _tgt_name, _src_name in _MS_MAP.items():
            _tgt = col_letters(_ms, _tgt_name)
            _src_col = _sg_headers.get(_src_name)
            if not _tgt or not _src_col:
                continue
            _src_letter = get_column_letter(_src_col)
            for _r in range(2, _ms.max_row + 1):
                _ms[f"{_tgt}{_r}"] = (
                    f'=IFERROR(INDEX(StageGates!${_src_letter}:${_src_letter},'
                    f'MATCH({_ms_key}{_r},StageGates!${_key_col_letter}:${_key_col_letter},0)),"")'
                )

        # Hide helper columns
        _sg.column_dimensions[_sg_key].hidden = True
        _ms.column_dimensions[_ms_key].hidden = True

# ── Financials ⟵ Projects (monthly CAPEX/OPEX baseline from Projects) ────
# CAPEX and OPEX per month = Projects totals / 12. Users can override any
# cell manually; Actual/Forecast/Variance stay editable.
if "Financials" in wb.sheetnames:
    _fin = wb["Financials"]
    _fin_pid = col_letters(_fin, "Project ID")
    _p_capex_idx = _proj_header_idx.get("CAPEX")
    _p_opex_idx  = _proj_header_idx.get("OPEX")
    if _fin_pid and _p_capex_idx and _p_opex_idx:
        _range = f"Projects!$A:${_proj_max_col}"
        for _hdr, _idx in (("CAPEX", _p_capex_idx), ("OPEX", _p_opex_idx)):
            _tgt = col_letters(_fin, _hdr)
            if not _tgt:
                continue
            for _r in range(2, _fin.max_row + 1):
                _fin[f"{_tgt}{_r}"] = (
                    f'=IFERROR(VLOOKUP({_fin_pid}{_r},{_range},{_idx},FALSE)/12,0)'
                )



# Governance: Next Gate via hidden lookup
stg = wb.create_sheet("_Stages")
stg.sheet_state = "hidden"
for i, s in enumerate(STAGES, start=1):
    stg.cell(row=i, column=1, value=s)
stage_range = f"_Stages!$A$1:$A${len(STAGES)}"
_gov_name = "Governance" if "Governance" in wb.sheetnames else (
    "StageGates" if "StageGates" in wb.sheetnames else None)
if _gov_name:
    ws = wb[_gov_name]
    cur_col = col_letters(ws, "Stage"); nxt_col = col_letters(ws, "Next Gate")
    if cur_col and nxt_col:
        for r in range(2, ws.max_row + 1):
            ws[f"{nxt_col}{r}"] = (
                f"=IFERROR(INDEX({stage_range},"
                f"MIN(MATCH({cur_col}{r},{stage_range},0)+1,{len(STAGES)})),\"\")"
            )

# ───────────────────────────── DATA VALIDATIONS (from Config) ────────────
from utils.excel_validations import apply_validations_from_config
apply_validations_from_config(wb)

# Header formatting
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill("solid", start_color="1F4E78")
for sname in wb.sheetnames:
    sh = wb[sname]
    if sh.sheet_state == "hidden":
        continue
    for cell in sh[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    sh.freeze_panes = "A2"

wb.save(OUT)

# ───────────────────────────── RECALCULATE ─────────────────────────────
try:
    import subprocess, shutil, tempfile
    soffice = shutil.which("soffice") or shutil.which("libreoffice")
    if soffice:
        with tempfile.TemporaryDirectory() as td:
            subprocess.run([soffice, "--headless", "--calc",
                            "--convert-to", "xlsx", "--outdir", td, str(OUT)],
                           check=True, timeout=180,
                           stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
            recalced = Path(td) / OUT.name
            if recalced.exists():
                shutil.copy(recalced, OUT)
                print("✅ Formulas recalculated via LibreOffice")
    else:
        print("⚠️  LibreOffice not found; open in Excel once to cache values.")
except Exception as e:
    print(f"⚠️  Recalc skipped: {e}")

print(f"✅ Wrote {OUT}")
