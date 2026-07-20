"""PMO Admin Console — control the app from one place.

Run:  streamlit run appadmin.py

Capabilities
------------
1. **Page Visibility** — toggle which pages appear in the main app's sidebar.
   Hidden pages are physically moved to `pages_disabled/` and restored on enable.
2. **Sheet Editor** — edit any sheet of the active workbook (full CRUD) and
   save back, with automatic timestamped backups.
3. **Data Source** — switch the active master Excel file or upload a new one.
4. **Backups** — list & restore any previous backup.
5. **Admin Auth** — simple password gate (default `admin` — change via
   `data/admin_config.json` → "password").
"""
from __future__ import annotations
import json
import shutil
from datetime import datetime
from pathlib import Path

import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

from config import (get_master_file, set_master_file, DEFAULT_MASTER_FILE,
                    UPLOAD_DIR, DATA_DIR)
from utils.excel_loader import refresh
from utils.theme_manager import apply_theme, theme_picker_sidebar, render_sheet

# ───────── config files ─────────
ROOT = Path(__file__).parent
PAGES_DIR = ROOT / "pages"
DISABLED_DIR = ROOT / "pages_disabled"
DISABLED_DIR.mkdir(exist_ok=True)
ADMIN_CFG = DATA_DIR / "admin_config.json"


def _load_admin_cfg() -> dict:
    if ADMIN_CFG.exists():
        try:
            return json.loads(ADMIN_CFG.read_text())
        except Exception:
            pass
    return {"password": "admin"}


def _save_admin_cfg(cfg: dict) -> None:
    ADMIN_CFG.write_text(json.dumps(cfg, indent=2))


cfg = _load_admin_cfg()

# ───────── page setup ─────────
st.set_page_config(page_title="PMO Admin Console", page_icon="🛡️", layout="wide")
apply_theme()
theme_picker_sidebar()

st.sidebar.title("🛡️ ADMIN CONSOLE")
st.sidebar.caption("PMO Enterprise — restricted area")

# ───────── auth ─────────
if "admin_ok" not in st.session_state:
    st.session_state.admin_ok = False

if not st.session_state.admin_ok:
    st.title("🛡️ PMO Admin Console")
    st.info("Enter the admin password to continue. Default: `admin` "
            "(change it under **Settings** once signed in).")
    pwd = st.text_input("Password", type="password")
    if st.button("Sign in", type="primary"):
        if pwd == cfg.get("password", "admin"):
            st.session_state.admin_ok = True
            st.rerun()
        else:
            st.error("Incorrect password.")
    st.stop()

st.title("🛡️ PMO Admin Console")
st.caption("Single place to manage page visibility, edit workbook data, "
           "switch data sources and restore backups.")

if st.sidebar.button("🔒 Sign out", use_container_width=True):
    st.session_state.admin_ok = False
    st.rerun()

tabs = st.tabs(["📑 Page Visibility", "🔀 Reorder Pages", "📝 Sheet Editor",
                "📂 Data Source", "🗄️ Backups", "⚙️ Settings"])

# ═══════════════════════════════════════════════════════════════════════
# 1) PAGE VISIBILITY
# ═══════════════════════════════════════════════════════════════════════
with tabs[0]:
    st.subheader("Control which pages appear in the main app")
    st.caption("Hidden pages are moved to `pages_disabled/` and restored on enable. "
               "The main app picks the change up on its next reload.")

    visible = sorted([p for p in PAGES_DIR.glob("*.py")])
    hidden  = sorted([p for p in DISABLED_DIR.glob("*.py")])
    all_pages = [(p.name, True) for p in visible] + [(p.name, False) for p in hidden]
    all_pages.sort(key=lambda x: x[0])

    if not all_pages:
        st.warning("No pages found under `pages/` or `pages_disabled/`.")
    else:
        c1, c2 = st.columns([3, 1])
        c1.markdown("**Page file**")
        c2.markdown("**Visible**")
        changes = {}
        for name, is_visible in all_pages:
            cc1, cc2 = st.columns([3, 1])
            label = name.replace(".py", "").replace("_", " ")
            cc1.markdown(f"`{name}` &nbsp; — &nbsp; {label}", unsafe_allow_html=True)
            new = cc2.toggle("", value=is_visible, key=f"vis_{name}",
                             label_visibility="collapsed")
            changes[name] = (is_visible, new)

        if st.button("💾 Apply visibility changes", type="primary"):
            moved = 0
            for name, (was, now) in changes.items():
                if was == now: continue
                src = (PAGES_DIR if was else DISABLED_DIR) / name
                dst = (DISABLED_DIR if was else PAGES_DIR) / name
                try:
                    shutil.move(str(src), str(dst))
                    moved += 1
                except Exception as e:
                    st.error(f"Failed to move {name}: {e}")
            if moved:
                st.success(f"Applied {moved} change(s). Reload the main app to see them.")
            else:
                st.info("No changes to apply.")

        with st.expander("ℹ️ How it works"):
            st.markdown(
                "- Streamlit auto-discovers files in `pages/` as sidebar entries.\n"
                "- Toggling **off** moves the file to `pages_disabled/` — it disappears "
                "from the sidebar but stays in the repo for re-enabling later.\n"
                "- Toggling **on** moves it back to `pages/`.\n"
                "- The main running app keeps its current navigation until reloaded."
            )

# ═══════════════════════════════════════════════════════════════════════
# 1b) REORDER PAGES
# ═══════════════════════════════════════════════════════════════════════
with tabs[1]:
    st.subheader("Reorder pages in the main app sidebar")
    st.caption("Streamlit orders pages by the numeric prefix in the filename. "
               "Drag the order below (use the number inputs) and apply — files "
               "will be renamed with new prefixes. Reload the main app to see changes.")

    import re as _re
    def _split_prefix(name: str):
        m = _re.match(r"^(\d+)[_\-\s]+(.*)$", name)
        if m:
            return int(m.group(1)), m.group(2)
        return 9999, name

    visible_files = sorted(PAGES_DIR.glob("*.py"), key=lambda p: _split_prefix(p.name))
    if not visible_files:
        st.info("No visible pages to reorder.")
    else:
        if "page_order" not in st.session_state or \
           set(st.session_state.page_order) != {p.name for p in visible_files}:
            st.session_state.page_order = [p.name for p in visible_files]

        order = list(st.session_state.page_order)
        st.markdown("**Current order** (top = first in sidebar)")
        for i, name in enumerate(order):
            c1, c2, c3, c4 = st.columns([6, 1, 1, 1])
            label = _split_prefix(name)[1].replace(".py", "").replace("_", " ")
            c1.markdown(f"**{i+1}.** `{name}` — {label}")
            if c2.button("⬆️", key=f"up_{name}", disabled=(i == 0)):
                order[i-1], order[i] = order[i], order[i-1]
                st.session_state.page_order = order
                st.rerun()
            if c3.button("⬇️", key=f"dn_{name}", disabled=(i == len(order)-1)):
                order[i+1], order[i] = order[i], order[i+1]
                st.session_state.page_order = order
                st.rerun()
            if c4.button("⏫ Top", key=f"top_{name}", disabled=(i == 0)):
                order.insert(0, order.pop(i))
                st.session_state.page_order = order
                st.rerun()

        st.markdown("---")
        cA, cB = st.columns([1, 1])
        if cA.button("💾 Apply new order", type="primary"):
            # rename in two passes to avoid collisions
            tmp_map = {}
            for i, name in enumerate(order, start=1):
                _, base = _split_prefix(name)
                base = base.lstrip("_- ")
                new_name = f"{i}_{base}"
                if new_name == name:
                    continue
                tmp = PAGES_DIR / f"__tmp_{i}_{name}"
                (PAGES_DIR / name).rename(tmp)
                tmp_map[tmp] = PAGES_DIR / new_name
            for tmp, final in tmp_map.items():
                tmp.rename(final)
            st.success(f"Renamed {len(tmp_map)} file(s). Reload the main app to see the new order.")
            st.session_state.pop("page_order", None)
            st.rerun()
        if cB.button("↩︎ Reset to current file order"):
            st.session_state.pop("page_order", None)
            st.rerun()

# ═══════════════════════════════════════════════════════════════════════
# 2) SHEET EDITOR  (full workbook CRUD)
# ═══════════════════════════════════════════════════════════════════════
# 2) SHEET EDITOR  (full workbook CRUD)
# ═══════════════════════════════════════════════════════════════════════
with tabs[2]:
    st.subheader("Edit any sheet of the active workbook")
    master = Path(get_master_file())
    st.info(f"**Active file:** `{master}`", icon="📂")

    if not master.exists():
        st.error("Master file missing. Set it under **Data Source**.")
    else:
        try:
            xls = pd.ExcelFile(master, engine="openpyxl")
            sheets = list(xls.sheet_names)
        except Exception as e:
            st.error(f"Could not open workbook: {e}")
            sheets = []

        if sheets:
            sheet_name = st.selectbox("Sheet", sheets, key="admin_sheet")
            try:
                original = pd.read_excel(xls, sheet_name=sheet_name)
            except Exception as e:
                st.error(f"Read failed: {e}")
                original = pd.DataFrame()

            # normalize cols
            cols, seen = [], {}
            for c in original.columns:
                n = str(c).strip() or "Unnamed"
                if n.startswith("Unnamed"): n = "Unnamed"
                if n in seen:
                    seen[n] += 1; n = f"{n}_{seen[n]}"
                else:
                    seen[n] = 0
                cols.append(n)
            original.columns = cols
            original = original.reset_index(drop=True)
            if original.empty and len(original.columns) == 0:
                original = pd.DataFrame({"Column1": [""]})

            st.markdown(f"**Rows:** {len(original):,} · **Cols:** {len(original.columns):,}")
            st.markdown("**Preview**")
            render_sheet(original, max_rows=150)

            def _disp(v):
                if pd.isna(v): return ""
                if isinstance(v, (pd.Timestamp, datetime)):
                    return v.strftime("%Y-%m-%d") if getattr(v, "time", lambda: None)() == datetime.min.time() else str(v)
                return str(v)

            editor_df = original.astype("object").apply(lambda c: c.map(_disp))
            st.markdown("**Editable (add / delete / modify rows)**")
            edited = st.data_editor(
                editor_df, use_container_width=True, num_rows="dynamic",
                height=520, key=f"admin_editor_{sheet_name}")

            c1, c2, _ = st.columns([1,1,4])
            do_save = c1.button("💾 Save sheet", type="primary", use_container_width=True)
            if c2.button("↩️ Discard", use_container_width=True):
                st.rerun()

            if do_save:
                # restore numeric/date types based on original
                out = edited.copy().replace("", pd.NA)
                for col in out.columns:
                    if col not in original.columns: continue
                    src = original[col]
                    try:
                        if pd.api.types.is_datetime64_any_dtype(src):
                            out[col] = pd.to_datetime(out[col], errors="coerce")
                        elif pd.api.types.is_integer_dtype(src):
                            out[col] = pd.to_numeric(out[col], errors="coerce").astype("Int64")
                        elif pd.api.types.is_float_dtype(src):
                            out[col] = pd.to_numeric(out[col], errors="coerce")
                    except Exception: pass
                # backup
                bdir = DATA_DIR / "backups"; bdir.mkdir(exist_ok=True)
                bkp = bdir / f"{master.stem}__{datetime.now():%Y%m%d_%H%M%S}{master.suffix}"
                shutil.copy2(master, bkp)
                # write
                try:
                    wb = load_workbook(master)
                    if sheet_name in wb.sheetnames:
                        idx = wb.sheetnames.index(sheet_name)
                        del wb[sheet_name]
                        ws = wb.create_sheet(sheet_name, index=idx)
                    else:
                        ws = wb.create_sheet(sheet_name)
                    to_save = out.where(pd.notna(out), "")
                    for row in dataframe_to_rows(to_save, index=False, header=True):
                        ws.append(row)
                    wb.save(master)
                    refresh()
                    st.success(f"Saved **{sheet_name}** ({len(edited):,} rows). Backup: `{bkp.name}`")
                except Exception as e:
                    st.error(f"Save failed: {e}")

# ═══════════════════════════════════════════════════════════════════════
# 3) DATA SOURCE
# ═══════════════════════════════════════════════════════════════════════
with tabs[3]:
    st.subheader("Active workbook")
    current = get_master_file()
    st.code(str(current))
    new_path = st.text_input("Switch to file path", value=str(current))
    if st.button("Use this path"):
        p = Path(new_path).expanduser()
        if p.exists() and p.suffix.lower() in {".xlsx", ".xlsm"}:
            set_master_file(p); refresh()
            st.success(f"Active file → {p}")
        else:
            st.error("Invalid path.")
    up = st.file_uploader("Or upload a new .xlsx", type=["xlsx","xlsm"])
    if up:
        dest = UPLOAD_DIR / up.name
        dest.write_bytes(up.getbuffer())
        set_master_file(dest); refresh()
        st.success(f"Uploaded and activated `{dest.name}`")
    if st.button("↩︎ Reset to bundled sample"):
        set_master_file(DEFAULT_MASTER_FILE); refresh()
        st.success("Reset to bundled sample.")

# ═══════════════════════════════════════════════════════════════════════
# 4) BACKUPS
# ═══════════════════════════════════════════════════════════════════════
with tabs[4]:
    st.subheader("Backups (data/backups/)")
    bdir = DATA_DIR / "backups"; bdir.mkdir(exist_ok=True)
    bks = sorted(bdir.glob("*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)
    if not bks:
        st.info("No backups yet.")
    else:
        for b in bks[:40]:
            c1, c2, c3, c4 = st.columns([4,2,1,1])
            c1.markdown(f"`{b.name}`")
            c2.caption(datetime.fromtimestamp(b.stat().st_mtime).strftime("%Y-%m-%d %H:%M:%S"))
            with open(b, "rb") as f:
                c3.download_button("⬇︎", f.read(), file_name=b.name, key=f"dl_{b.name}")
            if c4.button("↺ Restore", key=f"rs_{b.name}"):
                target = Path(get_master_file())
                shutil.copy2(b, target); refresh()
                st.success(f"Restored {b.name} → {target.name}")

# ═══════════════════════════════════════════════════════════════════════
# 5) SETTINGS
# ═══════════════════════════════════════════════════════════════════════
with tabs[5]:
    st.subheader("Admin password")
    st.caption("Stored in `data/admin_config.json`. Change immediately after first run.")
    new_pwd = st.text_input("New password", type="password")
    new_pwd2 = st.text_input("Repeat", type="password")
    if st.button("Update password", type="primary"):
        if not new_pwd:
            st.error("Password cannot be empty.")
        elif new_pwd != new_pwd2:
            st.error("Passwords do not match.")
        else:
            cfg["password"] = new_pwd
            _save_admin_cfg(cfg)
            st.success("Password updated.")
