"""Read/write helpers that persist sheets back to the active Excel workbook
without disturbing formulas in OTHER sheets.

Strategy: load with openpyxl, delete & rewrite ONLY the target sheet, save.
Other sheets (and their formulas) remain untouched.

Concurrency: every write acquires a cross-process file lock (.lock sidecar)
so two users hitting Save at the same moment cannot corrupt the workbook.
Falls back to a best-effort no-op lock if the `filelock` package is missing.
"""
from __future__ import annotations
from pathlib import Path
from datetime import datetime
from contextlib import contextmanager
import shutil
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

from config import get_master_file, DATA_DIR
from utils.excel_loader import refresh

BACKUP_DIR = DATA_DIR / "backups"
BACKUP_DIR.mkdir(parents=True, exist_ok=True)

# Cross-process lock — protects the shared master file when the app is
# hosted centrally and multiple users click Save simultaneously.
try:
    from filelock import FileLock, Timeout as _LockTimeout
    _HAS_FILELOCK = True
except Exception:  # filelock not installed — degrade gracefully.
    _HAS_FILELOCK = False
    _LockTimeout = Exception  # type: ignore


@contextmanager
def _workbook_lock(path: Path, timeout: int = 30):
    """Acquire an exclusive lock on `<path>.lock` for the duration of a write."""
    if not _HAS_FILELOCK:
        yield
        return
    lock = FileLock(str(path) + ".lock", timeout=timeout)
    try:
        with lock:
            yield
    except _LockTimeout as e:
        raise RuntimeError(
            f"Another user is currently saving the workbook. "
            f"Please wait a few seconds and try again. ({e})"
        )


def write_sheet(sheet_name: str, df: pd.DataFrame) -> Path:
    """Replace `sheet_name` in the active workbook with `df`. Creates a
    timestamped backup and clears the in-process cache.

    If the workbook can't be opened by openpyxl (corrupted, truncated,
    locked-and-copied mid-write, legacy .xls, or a stale OneDrive
    placeholder), fall back to rebuilding every sheet from the in-memory
    cache with pandas so the save still succeeds.
    """
    path = get_master_file()
    if not path.exists():
        raise FileNotFoundError(path)

    with _workbook_lock(path):
        backup = BACKUP_DIR / f"{path.stem}_{datetime.now():%Y%m%d_%H%M%S}{path.suffix}"
        shutil.copy(path, backup)

        try:
            _write_sheet_openpyxl(path, sheet_name, df)
        except Exception as e:
            msg = str(e).lower()
            if ("content_types" in msg or "badzipfile" in msg or "not a zip" in msg
                    or "not a valid" in msg or "no item named" in msg
                    or isinstance(e, KeyError)):
                _write_sheet_rebuild(path, sheet_name, df)
            else:
                raise

    refresh()
    try:
        from utils.config_loader import refresh as cfg_refresh
        cfg_refresh()
    except Exception:
        pass
    return path


def _safe(v):
    if v is None:
        return None
    if isinstance(v, float) and pd.isna(v):
        return None
    if isinstance(v, pd.Timestamp):
        return v.to_pydatetime() if not pd.isna(v) else None
    return v


def _write_sheet_openpyxl(path: Path, sheet_name: str, df: pd.DataFrame) -> None:
    wb = load_workbook(path)
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    cols = [str(c) for c in df.columns]
    ws.append(cols)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", start_color="1F4E78")
    for cell in ws[1]:
        cell.font = header_font; cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"

    for _, row in df.iterrows():
        ws.append([_safe(v) for v in row.tolist()])

    try:
        from utils.excel_validations import apply_validations_from_config
        apply_validations_from_config(wb)
    except Exception:
        pass

    wb.save(path)


def _write_sheet_rebuild(path: Path, sheet_name: str, df: pd.DataFrame) -> None:
    """Rebuild the workbook from cached sheets + the new df when the on-disk
    file can't be opened by openpyxl. Formulas in other sheets materialise
    to their last cached values — better than the save failing outright."""
    from utils.excel_loader import load_all
    cached = load_all().get("_sheets", {}) or {}
    sheets: dict[str, pd.DataFrame] = {}
    for name, sdf in cached.items():
        if name == sheet_name:
            continue
        if isinstance(sdf, pd.DataFrame):
            sheets[name] = sdf
    sheets[sheet_name] = df
    with pd.ExcelWriter(path, engine="openpyxl", mode="w") as xw:
        for name, sdf in sheets.items():
            safe = sdf.copy()
            for c in safe.columns:
                if safe[c].dtype == "object":
                    safe[c] = safe[c].where(pd.notna(safe[c]), "")
            safe.to_excel(xw, sheet_name=str(name)[:31], index=False)


def append_row(sheet_name: str, row: dict) -> Path:
    """Append a single row to a sheet (preserves existing rows & headers)."""
    from utils.excel_loader import load_all
    data = load_all()
    df = data["_sheets"].get(sheet_name, pd.DataFrame())
    if df.empty:
        df = pd.DataFrame([row])
    else:
        # Add any new keys as new columns
        for k in row:
            if k not in df.columns:
                df[k] = None
        df = pd.concat([df, pd.DataFrame([row])[df.columns]], ignore_index=True)
    return write_sheet(sheet_name, df)


def update_row(sheet_name: str, key_col: str, key_val, updates: dict) -> Path:
    from utils.excel_loader import load_all
    df = load_all()["_sheets"].get(sheet_name, pd.DataFrame()).copy()
    if df.empty or key_col not in df.columns:
        raise ValueError(f"{key_col} not found in {sheet_name}")
    mask = df[key_col].astype(str) == str(key_val)
    for k, v in updates.items():
        if k not in df.columns:
            df[k] = None
        df.loc[mask, k] = v
    return write_sheet(sheet_name, df)


def delete_row(sheet_name: str, key_col: str, key_val) -> Path:
    from utils.excel_loader import load_all
    df = load_all()["_sheets"].get(sheet_name, pd.DataFrame()).copy()
    if df.empty:
        return get_master_file()
    df = df[df[key_col].astype(str) != str(key_val)]
    return write_sheet(sheet_name, df)
