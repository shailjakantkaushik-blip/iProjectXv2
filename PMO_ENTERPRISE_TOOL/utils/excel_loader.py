"""Load PMO data from the active master Excel workbook with caching.

Robust to schema changes: any extra columns you add in Excel flow through
automatically, and any sheet present in the workbook is exposed under
`data["_sheets"][<exact-sheet-name>]` so new sheets work without code
changes. Sheet-name matching for the canonical keys (projects, risks, …)
is case/whitespace-insensitive.
"""
from __future__ import annotations
import pandas as pd
import numpy as np
import streamlit as st
from config import get_master_file, SHEETS


def _norm(s: str) -> str:
    return "".join(str(s).lower().split())


# Expected columns per canonical sheet and the default value to inject when
# the column is absent in the user's workbook. "0" → numeric zero, "NA" →
# pandas NA, "date" → NaT, "" → empty string.
EXPECTED: dict[str, dict[str, str]] = {
    "projects": {
        "Project ID":"NA","Project Name":"NA","Program":"NA","Sponsor":"NA",
        "Portfolio Category":"NA","Funding Type":"NA","Governance Channel":"NA",
        "Priority":"NA","Status":"NA","RAG":"NA","Current Phase":"NA",
        "Gate Status":"NA","Next Gate":"NA","Checklist Complete %":"0",
        "Delivery Method":"NA",
        "Start Date":"date","End Date":"date","Target Go-Live":"date","Target Date":"date","Go Live Date":"date",
        "Budget":"0","CAPEX Approved":"0","CAPEX Incurred":"0",
        "OPEX Approved":"0","OPEX Incurred":"0","ROI %":"0","NPV":"0",
        "Benefits Realised":"0","Schedule RAG":"NA","Financial RAG":"NA",
        "Delivery RAG":"NA",
    },
    "risks":        {"Risk ID":"NA","Project":"NA","Description":"","Severity":"NA","Probability":"NA","Impact":"NA","Owner":"NA","Status":"NA","Mitigation":""},
    "raid":         {"ID":"NA","Project":"NA","Type":"NA","Description":"","Owner":"NA","Status":"NA","Due Date":"date"},
    # Governance + StageGates are merged into one sheet ("StageGates") with the
    # union of both column sets. Legacy aliases (Target Date / Actual Date /
    # Checklist %) are mirrored to/from the canonical names after coerce.
    "governance":   {"Project ID":"NA","Project Name":"NA","Project":"NA","Stage":"NA","Status":"NA","Gate Status":"NA","Next Gate":"NA","Checklist Complete %":"0","Gate Owner":"NA","Planned Gate Date":"date","Actual Gate Date":"date","Gate Outcome":"NA","Gate Comments":"","Governance Channel":"NA","Channel":"NA","Days Late":"0","Owner":"NA","Date":"date","Decision":"","Target Date":"date","Actual Date":"date","Checklist %":"0"},
    "stagegates":   {"Project ID":"NA","Project Name":"NA","Project":"NA","Stage":"NA","Status":"NA","Gate Status":"NA","Next Gate":"NA","Checklist Complete %":"0","Gate Owner":"NA","Planned Gate Date":"date","Actual Gate Date":"date","Gate Outcome":"NA","Gate Comments":"","Governance Channel":"NA","Channel":"NA","Days Late":"0","Owner":"NA","Date":"date","Decision":"","Target Date":"date","Actual Date":"date","Checklist %":"0"},
    "financials":   {"Project":"NA","Year":"0","CAPEX":"0","OPEX":"0","Benefit":"0","Variance":"0"},
    "resources":    {"Resource":"NA","Role":"NA","Project":"NA","Allocation %":"0","Capacity %":"0","Month":"NA"},
    "dependencies": {"From Project":"NA","To Project":"NA","Dependency Type":"NA","Status":"NA","Impact":"NA"},
    "pipeline":     {"Idea":"NA","Sponsor":"NA","Stage":"NA","Estimated Value":"0","Status":"NA"},
    "costbenefit":  {"Project":"NA","Year":"0","CAPEX":"0","OPEX":"0","Benefit Type":"NA","Benefit":"0"},
    "benefits":     {"Project":"NA","Benefit":"","Type":"NA","Target":"0","Realised":"0","Owner":"NA"},
    "decisions":    {"Decision ID":"NA","Project":"NA","Decision":"","Date":"date","Owner":"NA","Status":"NA"},
    "actions":      {"Action ID":"NA","Project":"NA","Action":"","Owner":"NA","Due Date":"date","Status":"NA"},
    "milestones":   {"Project":"NA","Milestone":"NA","Date":"date","Status":"NA"},
    "portfoliomovements": {"Project":"NA","From":"NA","To":"NA","Date":"date","Reason":""},
    "prioritisation":     {"Project":"NA","Score":"0","Rank":"0","Criteria":"NA"},
    "fyallocation":       {"Project ID":"NA","Project Name":"NA","FY":"NA",
                           "Budget %":"0","Forecast %":"0",
                           "Budget Amount":"0","Forecast Amount":"0","Notes":""},
    "programs":           {"Program":"NA","Owner":"NA","Sponsor":"NA",
                           "Budget":"0","Forecast":"0","Start FY":"NA","End FY":"NA",
                           "Status":"NA","Notes":""},
    "projectbrief":       {"Project ID":"NA",
                           "Portfolio / Workstream":"","Sponsor":"","Business Owner":"",
                           "Business Solution Manager":"","Strategic Alignment":"",
                           "Background and Context":"","Opportunity / Problem Statement":"",
                           "Objective":"","In Scope":"","Out of Scope":"",
                           "Assumptions & Constraints":"","Key Metrics / Success Measures":"",
                           "Approval Type":"","Funding Ask":"","Funding Source":"",
                           "Resource Ask":"","Estimate Commentary":"",
                           "P&L Benefits Commentary":"","Delivery Milestones":"",
                           "Project Risks":"","Dependencies":""},
    "projectlinks":       {"Project ID":"NA","Title":"","URL":"","Category":"","Added":"date"},
    "phasefinancials":    {"Project ID":"NA","Stage":"NA",
                           "Planned Start":"date","Planned End":"date",
                           "Actual Start":"date","Actual End":"date",
                           "Phase Budget":"0","Phase Forecast":"0","Phase Actual Spend":"0",
                           "Status":"NA","Notes":""},
    "sprints":            {"Sprint ID":"NA","Project ID":"NA","Project Name":"NA",
                           "Sprint #":"0","Sprint Name":"NA",
                           "Start Date":"date","End Date":"date",
                           "Points Committed":"0","Points Completed":"0",
                           "Stories Committed":"0","Stories Completed":"0",
                           "Status":"NA","Team":"NA"},
}



def _coerce(df: pd.DataFrame, key: str) -> pd.DataFrame:
    spec = EXPECTED.get(key)
    if spec is None or df is None:
        return df if df is not None else pd.DataFrame()
    for col, kind in spec.items():
        if col not in df.columns:
            if kind == "0":      df[col] = 0
            elif kind == "date": df[col] = pd.NaT
            elif kind == "":     df[col] = ""
            else:                df[col] = pd.NA
        else:
            # Fill missing values inside an existing column with sensible defaults.
            if kind == "0":
                df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)
            elif kind == "date":
                df[col] = pd.to_datetime(df[col], errors="coerce")
            elif kind == "":
                df[col] = df[col].fillna("")
            else:
                df[col] = df[col].fillna("NA")
    return df


def _mtime(p: str) -> float:
    try:
        import os
        return os.path.getmtime(p)
    except Exception:
        return 0.0


@st.cache_data(show_spinner=False, max_entries=8)
def _load_cached(path: str, _mtime_key: float) -> dict:
    p = path
    # Fast path: openpyxl read_only + data_only skips formula parsing and
    # streams rows instead of building the full in-memory workbook. Roughly
    # 3–5× faster on the PMO master file. Fall back to the default engine
    # if anything goes wrong (e.g. legacy .xls, corrupt zip).
    raw: dict[str, pd.DataFrame] = {}
    try:
        from openpyxl import load_workbook
        wb = load_workbook(p, read_only=True, data_only=True)
        for s in wb.sheetnames:
            try:
                ws = wb[s]
                rows = ws.iter_rows(values_only=True)
                header = next(rows, None)
                if not header:
                    raw[s] = pd.DataFrame(); continue
                cols = [str(c).strip() if c is not None else "" for c in header]
                data_rows = list(rows)
                df = pd.DataFrame(data_rows, columns=cols)
                raw[s] = df
            except Exception:
                raw[s] = pd.DataFrame()
        wb.close()
    except Exception:
        try:
            xls = pd.ExcelFile(p, engine="openpyxl")
            for s in xls.sheet_names:
                try:
                    df = pd.read_excel(xls, sheet_name=s)
                    df.columns = [str(c).strip() for c in df.columns]
                    raw[s] = df
                except Exception:
                    raw[s] = pd.DataFrame()
        except Exception:
            raw = {}

    norm_map = {_norm(s): s for s in raw}

    # --- Merge legacy Governance + StageGates into one unified DataFrame ---
    # Older workbooks shipped both sheets. New workbooks ship only "StageGates".
    # If both exist, union their rows (Governance columns win where names clash);
    # then expose the combined frame under both the "governance" and "stagegates"
    # keys so all existing pages keep working.
    gov_raw = raw.get(norm_map.get(_norm("Governance"), ""), pd.DataFrame())
    sg_raw  = raw.get(norm_map.get(_norm("StageGates"),  ""), pd.DataFrame())
    if not gov_raw.empty and not sg_raw.empty:
        unified = pd.concat([gov_raw, sg_raw], ignore_index=True, sort=False)
    elif not gov_raw.empty:
        unified = gov_raw.copy()
    else:
        unified = sg_raw.copy()

    # Mirror alias columns so downstream code that reads either name works.
    _ALIASES = [
        ("Planned Gate Date", "Target Date"),
        ("Actual Gate Date",  "Actual Date"),
        ("Checklist Complete %", "Checklist %"),
        ("Gate Owner", "Owner"),
        ("Governance Channel", "Channel"),
    ]
    for canon, alias in _ALIASES:
        if canon in unified.columns and alias not in unified.columns:
            unified[alias] = unified[canon]
        elif alias in unified.columns and canon not in unified.columns:
            unified[canon] = unified[alias]

    data: dict = {}
    for key, sheet in SHEETS.items():
        if key in ("governance", "stagegates"):
            data[key] = _coerce(unified.copy(), key)
            continue
        match = norm_map.get(_norm(sheet))
        df = raw[match].copy() if match else pd.DataFrame()
        data[key] = _coerce(df, key)

    data["_sheets"] = raw
    return data


def _blank(v) -> bool:
    if v is None: return True
    try:
        if pd.isna(v): return True
    except Exception: pass
    s = str(v).strip()
    return s == "" or s.lower() in ("nan", "nat", "none")


def _is_formula(v) -> bool:
    return isinstance(v, str) and v.startswith("=")


def _num(v):
    try:
        f = float(v)
        return f if pd.notna(f) else 0.0
    except Exception:
        return 0.0


def _resolve_derived_formulas(data: dict) -> dict:
    """Derived sheets store live Excel formulas (Project Name via INDEX/MATCH,
    Total Cost = CAPEX+OPEX, etc.). openpyxl can't evaluate formulas without
    a cached value from Excel, so the app would see None. Recompute those
    columns in-memory so every chart/KPI works even on a freshly rebuilt
    workbook that Excel hasn't opened yet."""
    sheets = data.get("_sheets", {}) or {}
    projects = data.get("projects", pd.DataFrame())

    # Build Project ID → Name / CAPEX / OPEX / Approved Funding map once.
    # These fallbacks let CostBenefit / Financials compute Total Cost when the
    # per-row CAPEX/OPEX cells are blank or contain Excel formulas that
    # openpyxl can't evaluate (common on freshly rebuilt workbooks).
    name_by_id: dict[str, str] = {}
    capex_by_id: dict[str, float] = {}
    opex_by_id:  dict[str, float] = {}
    if not projects.empty and "Project ID" in projects.columns:
        for _, r in projects.iterrows():
            pid = r.get("Project ID")
            if _blank(pid): continue
            key = str(pid).strip()
            name_by_id[key]  = str(r.get("Project Name") or "").strip()
            capex_by_id[key] = _num(r.get("CAPEX")) or _num(r.get("Approved Funding"))
            opex_by_id[key]  = _num(r.get("OPEX"))

    def _tb(r):
        rec = r.get("Benefit Recurring"); one = r.get("Benefit One-Off")
        if not _blank(rec) or not _blank(one):
            return _num(rec) + _num(one)
        return _num(r.get("Benefit"))

    def _tc(r):
        c, o = r.get("CAPEX"), r.get("OPEX")
        cn, on = _num(c), _num(o)
        if cn > 0 or on > 0:
            return cn + on
        # CAPEX/OPEX were blank / formula / zero → fall back to Projects total
        # split evenly over the 5-year CostBenefit horizon.
        pid = str(r.get("Project ID") or "").strip()
        if pid and pid in capex_by_id:
            return (capex_by_id.get(pid, 0.0) + opex_by_id.get(pid, 0.0)) / 5.0
        return 0.0

    ARITH = {
        "CostBenefit": {
            "Total Cost":    _tc,
            "Total Benefit": _tb,
            "Net Benefit":   lambda r: _tb(r) - _tc(r),
        },
        "Financials": {
            "Actual":        _tc,
            "Forecast":      _tc,
            "Planned Value": _tc,
            "Earned Value":  _tc,
            "Variance":      lambda r: 0.0,
        },
    }

    # Columns we should ADD when missing (so charts have something to group on)
    ENSURE = {
        "CostBenefit": ["Total Cost", "Total Benefit", "Net Benefit"],
        "Financials":  [],
    }

    def _fix(df: pd.DataFrame, sheet_name: str) -> pd.DataFrame:
        if df is None or df.empty:
            return df
        df = df.copy()
        # 1) Project Name lookup
        if "Project Name" in df.columns and "Project ID" in df.columns:
            def _resolve_name(row):
                v = row.get("Project Name")
                if _is_formula(v) or _blank(v):
                    return name_by_id.get(str(row.get("Project ID") or "").strip(), "")
                return v
            df["Project Name"] = df.apply(_resolve_name, axis=1)
        # 2) Ensure derived numeric columns exist
        for col in ENSURE.get(sheet_name, []):
            if col not in df.columns:
                df[col] = np.nan
        # 3) Arithmetic formulas — recompute where formula/blank
        rules = ARITH.get(sheet_name, {})
        for col, fn in rules.items():
            if col not in df.columns: continue
            need_fix = df[col].apply(lambda v: _is_formula(v) or _blank(v))
            if need_fix.any():
                df.loc[need_fix, col] = df[need_fix].apply(fn, axis=1)
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)
        return df

    KEY_TO_SHEET = {"costbenefit": "CostBenefit", "financials": "Financials"}
    for key, sheet_name in KEY_TO_SHEET.items():
        if key in data:
            data[key] = _fix(data[key], sheet_name)

    for sheet_name in list(sheets.keys()):
        arith_key = sheet_name if sheet_name in ARITH else ""
        sheets[sheet_name] = _fix(sheets[sheet_name], arith_key)

    data["_sheets"] = sheets
    return data


def load_all(path_str: str | None = None) -> dict:
    p = path_str or str(get_master_file())
    data = _load_cached(p, _mtime(p))
    data = _resolve_derived_formulas(dict(data))
    try:
        from utils.data_sync import hydrate
        data = hydrate(data)
    except Exception:
        pass
    return data



def refresh():
    _load_cached.clear()
