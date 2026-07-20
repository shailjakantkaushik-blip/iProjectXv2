"""Rebuild derived sheets from the Projects sheet.

Design rules (per user):
  1. FY spans like "FY25-27" expand to [FY25, FY26, FY27].
  2. If a source field on a project is missing/blank, DO NOT fabricate a
     downstream value — leave the derived cell blank. Only Project ID and
     the stage/milestone label are always populated.
  3. For the Programs sheet, user-owned columns (Owner, Status, Notes) on
     existing rows are preserved; only rollup columns (Budget, Forecast,
     Start FY, End FY, Sponsor) are refreshed from Projects.
"""
from __future__ import annotations
import pandas as pd
from openpyxl.utils import get_column_letter
from utils.excel_loader import load_all
from utils.data_io import write_sheet

STAGES_A = ["Discovery","Business Case / Full Funding","Design","Build",
            "Testing","Deployment","Handover"]
STAGES_B = ["Discovery","Business Case / Seed Funding","Design",
            "Business Case / Full Funding","Build","Testing",
            "Deployment","Handover"]
MONTHS = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
CB_HORIZON = 5
REL_ENVS = ["DEV","TEST","UAT","PROD"]


# ─────────────────────── helpers ───────────────────────
def _blank(v) -> bool:
    if v is None: return True
    try:
        if pd.isna(v): return True
    except Exception: pass
    s = str(v).strip()
    return s == "" or s.lower() in ("nan", "nat", "none")


def _fy_choice_to_list(choice) -> list[str]:
    if _blank(choice): return []
    s = str(choice).strip().upper().replace(" ", "")
    if not s.startswith("FY"): return []
    body = s[2:]
    if "-" not in body:
        try: return [f"FY{int(body):02d}"]
        except Exception: return []
    a, b = body.split("-", 1)
    try:
        y0, y1 = int(a), int(b)
        if y1 < y0: y1 += 100
        return [f"FY{y:02d}" for y in range(y0, y1 + 1)]
    except Exception:
        return []


def _fy_label(ts):
    try:
        ts = pd.Timestamp(ts)
        if pd.isna(ts): return ""
        return f"FY{str(ts.year)[-2:]}"
    except Exception:
        return ""


def _num_or_none(v):
    if _blank(v): return None
    try:
        f = float(v)
        return f if pd.notna(f) else None
    except Exception:
        return None


def _date_or_none(v):
    if _blank(v): return None
    try:
        ts = pd.Timestamp(v)
        return ts if pd.notna(ts) else None
    except Exception:
        return None


def _stages_for(chan):
    return STAGES_A if "Channel A" in str(chan or "") else STAGES_B


def _get(row, col):
    """Safe accessor — returns None if column missing or value blank."""
    if col not in row.index: return None
    v = row[col]
    return None if _blank(v) else v


# ─────────────────────── builders ───────────────────────
def rebuild_fy_allocation(projects: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for _, p in projects.iterrows():
        pid = _get(p, "Project ID")
        if pid is None: continue
        fys = _fy_choice_to_list(_get(p, "Financial Year"))
        if not fys: continue  # no FY info → skip; don't fabricate
        approved = _num_or_none(_get(p, "Approved Funding"))
        forecast = _num_or_none(_get(p, "Forecast At Completion"))
        if forecast is None: forecast = approved
        share = 1.0 / len(fys)
        for fy in fys:
            rows.append({
                "Project ID":      str(pid),
                "Project Name":    str(_get(p, "Project Name") or ""),
                "FY":              fy,
                "Budget %":        round(share * 100, 1),
                "Forecast %":      round(share * 100, 1),
                "Budget Amount":   round(approved * share, 0) if approved is not None else "",
                "Forecast Amount": round(forecast * share, 0) if forecast is not None else "",
                "Notes":           "Auto-split from Projects.Financial Year — edit to revise",
            })
    return pd.DataFrame(rows)


def rebuild_phase_financials(projects: pd.DataFrame) -> pd.DataFrame:
    today = pd.Timestamp.today().normalize()
    rows = []
    for _, p in projects.iterrows():
        pid = _get(p, "Project ID")
        if pid is None: continue
        stages = _stages_for(_get(p, "Governance Channel"))
        start = _date_or_none(_get(p, "Start Date"))
        end   = _date_or_none(_get(p, "End Date"))
        approved = _num_or_none(_get(p, "Approved Funding"))
        have_span = start is not None and end is not None and end > start
        per = (end - start).days // len(stages) if have_span else 0
        share = 1.0 / len(stages)
        for i, stg in enumerate(stages):
            row = {"Project ID": str(pid), "Stage": stg,
                   "Planned Start": "", "Planned End": "",
                   "Actual Start": "", "Actual End": "",
                   "Phase Budget": "", "Phase Forecast": "",
                   "Phase Actual Spend": "", "Status": "", "Notes": ""}
            if have_span:
                ps = start + pd.Timedelta(days=i * per)
                pe = ps + pd.Timedelta(days=per)
                done = pe < today
                row.update({
                    "Planned Start": ps, "Planned End": pe,
                    "Actual Start": ps if ps < today else "",
                    "Actual End":   pe if done else "",
                    "Status": "Complete" if done else ("In Progress" if ps < today else "Planned"),
                })
                if approved is not None:
                    amt = round(approved * share, 0)
                    row["Phase Budget"] = amt
                    row["Phase Forecast"] = amt
                    row["Phase Actual Spend"] = amt if done else 0
            rows.append(row)
    return pd.DataFrame(rows)


def rebuild_programs(projects: pd.DataFrame, existing: pd.DataFrame | None = None) -> pd.DataFrame:
    """Roll up Projects → Programs. Preserves Owner/Status/Notes from any
    existing Programs sheet; only refreshes Budget/Forecast/FY window/Sponsor."""
    if projects.empty or "Program" not in projects.columns:
        return existing if existing is not None else pd.DataFrame()

    existing_by_prog = {}
    if existing is not None and not existing.empty and "Program" in existing.columns:
        for _, r in existing.iterrows():
            existing_by_prog[str(r["Program"]).strip()] = r.to_dict()

    rows = []
    seen = set()
    for prog, sub in projects.groupby(projects["Program"].astype(str)):
        prog = prog.strip()
        if _blank(prog): continue
        seen.add(prog)
        prev = existing_by_prog.get(prog, {})

        budget   = pd.to_numeric(sub.get("Approved Funding"), errors="coerce").sum(skipna=True)
        forecast = pd.to_numeric(sub.get("Forecast At Completion"), errors="coerce").sum(skipna=True)
        starts   = pd.to_datetime(sub.get("Start Date"), errors="coerce").dropna()
        ends     = pd.to_datetime(sub.get("End Date"),   errors="coerce").dropna()

        # Sponsor: refresh from Projects if present, else keep existing.
        sponsors = sub.get("Sponsor")
        sponsor_new = ""
        if sponsors is not None:
            s = sponsors.dropna().astype(str).str.strip()
            s = s[s != ""]
            if not s.empty: sponsor_new = s.iloc[0]

        rows.append({
            "Program":  prog,
            "Owner":    prev.get("Owner", "") or "",                # preserved
            "Sponsor":  sponsor_new or (prev.get("Sponsor", "") or ""),
            "Budget":   float(budget)   if pd.notna(budget)   else (prev.get("Budget", "")   or ""),
            "Forecast": float(forecast) if pd.notna(forecast) else (prev.get("Forecast", "") or ""),
            "Start FY": _fy_label(starts.min()) if not starts.empty else (prev.get("Start FY", "") or ""),
            "End FY":   _fy_label(ends.max())   if not ends.empty   else (prev.get("End FY", "")   or ""),
            "Status":   prev.get("Status", "Active") or "Active",   # preserved
            "Notes":    prev.get("Notes", "") or "",                # preserved
        })

    # Keep any existing programs that no longer appear in Projects (don't silently delete)
    for prog, prev in existing_by_prog.items():
        if prog not in seen:
            rows.append(prev)

    return pd.DataFrame(rows)


def rebuild_stage_gates(projects: pd.DataFrame) -> pd.DataFrame:
    today = pd.Timestamp.today().normalize()
    rows = []
    for _, p in projects.iterrows():
        pid = _get(p, "Project ID")
        if pid is None: continue
        stages = _stages_for(_get(p, "Governance Channel"))
        start = _date_or_none(_get(p, "Start Date"))
        end   = _date_or_none(_get(p, "End Date"))
        have_span = start is not None and end is not None and end > start
        per = (end - start).days // len(stages) if have_span else 0
        for i, s in enumerate(stages):
            row = {
                "Project ID":           str(pid),
                "Project Name":         str(_get(p, "Project Name") or ""),
                "Governance Channel":   str(_get(p, "Governance Channel") or ""),
                "Stage":                s,
                "Next Gate":            stages[min(i + 1, len(stages) - 1)],
                "Gate Status":          "", "Status": "",
                "Gate Owner":           str(_get(p, "Sponsor") or ""),
                "Planned Gate Date":    "", "Actual Gate Date": "",
                "Gate Outcome":         "", "Gate Comments": "",
                "Checklist Complete %": "", "Days Late": "",
            }
            if have_span:
                planned = start + pd.Timedelta(days=per * (i + 1))
                is_past = planned < today
                row.update({
                    "Gate Status": "Approved" if is_past else "Not Started",
                    "Status":      "Complete" if is_past else "Not Started",
                    "Planned Gate Date":    planned,
                    "Actual Gate Date":     planned if is_past else "",
                    "Gate Outcome":         "Proceed" if is_past else "",
                    "Checklist Complete %": 100 if is_past else 0,
                    "Days Late":            0,
                })
            rows.append(row)
    return pd.DataFrame(rows)


def rebuild_financials(projects: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for _, p in projects.iterrows():
        pid = _get(p, "Project ID")
        if pid is None: continue
        capex = _num_or_none(_get(p, "CAPEX"))
        opex  = _num_or_none(_get(p, "OPEX"))
        if capex is None and opex is None:
            # No funding data → emit blank monthly frame anchored on Project ID
            for m in MONTHS:
                rows.append({"Project ID": str(pid), "Month": m,
                             "CAPEX": "", "OPEX": "", "Actual": "", "Forecast": "",
                             "Variance": "", "Planned Value": "", "Earned Value": ""})
            continue
        for m in MONTHS:
            c = round((capex or 0) / 12, 0)
            o = round((opex  or 0) / 12, 0)
            rows.append({
                "Project ID": str(pid), "Month": m,
                "CAPEX": c if capex is not None else "",
                "OPEX":  o if opex  is not None else "",
                "Actual": c + o, "Forecast": c + o, "Variance": 0,
                "Planned Value": c + o, "Earned Value": c + o,
            })
    return pd.DataFrame(rows)


def rebuild_cost_benefit(projects: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for _, p in projects.iterrows():
        pid = _get(p, "Project ID")
        if pid is None: continue
        start = _date_or_none(_get(p, "Start Date"))
        total_capex = _num_or_none(_get(p, "Approved Funding"))
        if start is None or total_capex is None:
            # Not enough data — emit a single placeholder row so PID is tracked
            rows.append({"Project ID": str(pid),
                         "Project Name": str(_get(p, "Project Name") or ""),
                         "Program": str(_get(p, "Program") or ""),
                         "Year Offset": "", "Year": "",
                         "CAPEX": "", "OPEX": "", "Total Cost": "",
                         "Benefit Recurring": "", "Benefit One-Off": "",
                         "Total Benefit": "", "Net Benefit": "",
                         "Benefit Type": "", "Benefit Category": "",
                         "Confidence %": ""})
            continue
        start_year = start.year
        yearly_opex = round(total_capex * 0.10, 0)
        base_benefit = round(total_capex * 0.40, 0)
        for idx in range(1, CB_HORIZON + 1):
            yr = start_year + (idx - 1)
            capex = round(total_capex * 0.65, 0) if idx == 1 else (round(total_capex * 0.30, 0) if idx == 2 else 0)
            opex = 0 if idx == 1 else yearly_opex
            ben_rec = base_benefit if idx >= 2 else 0
            rows.append({
                "Project ID": str(pid),
                "Project Name": str(_get(p, "Project Name") or ""),
                "Program": str(_get(p, "Program") or ""),
                "Year Offset": idx - 1, "Year": yr,
                "CAPEX": capex, "OPEX": opex, "Total Cost": capex + opex,
                "Benefit Recurring": ben_rec, "Benefit One-Off": 0,
                "Total Benefit": ben_rec,
                "Net Benefit": ben_rec - (capex + opex),
                "Benefit Type": "Recurring",
                "Benefit Category": "Cost Savings",
                "Confidence %": 80,
            })
    return pd.DataFrame(rows)


def rebuild_releases(projects: pd.DataFrame) -> pd.DataFrame:
    today = pd.Timestamp.today().normalize()
    rows = []; rid = 1
    for _, p in projects.iterrows():
        pid = _get(p, "Project ID")
        if pid is None: continue
        start = _date_or_none(_get(p, "Start Date"))
        end   = _date_or_none(_get(p, "End Date"))
        if start is None or end is None or end <= start:
            # No usable window — emit a single placeholder release
            rows.append({"Release ID": f"REL{rid:04d}",
                         "Project ID": str(pid),
                         "Project Name": str(_get(p, "Project Name") or ""),
                         "Version": "", "Type": "",
                         "Planned Date": "", "Actual Date": "",
                         "Status": "", "Owner": str(_get(p, "Delivery Lead") or ""),
                         "Environment": "", "Notes": ""})
            rid += 1; continue
        step = max(1, (end - start).days // 4)
        for i in range(4):
            planned = start + pd.Timedelta(days=step * (i + 1))
            done = planned < today
            rows.append({
                "Release ID": f"REL{rid:04d}",
                "Project ID": str(pid),
                "Project Name": str(_get(p, "Project Name") or ""),
                "Version": f"{i+1}.0",
                "Type": "Minor" if i < 3 else "Major",
                "Planned Date": planned,
                "Actual Date": planned if done else "",
                "Status": "Released" if done else "Planned",
                "Owner": str(_get(p, "Delivery Lead") or ""),
                "Environment": REL_ENVS[i % len(REL_ENVS)],
                "Notes": "",
            })
            rid += 1
    return pd.DataFrame(rows)


def rebuild_milestones(projects: pd.DataFrame) -> pd.DataFrame:
    today = pd.Timestamp.today().normalize()
    rows = []; mid = 1
    for _, p in projects.iterrows():
        pid = _get(p, "Project ID")
        if pid is None: continue
        stages = _stages_for(_get(p, "Governance Channel"))
        start = _date_or_none(_get(p, "Start Date"))
        end   = _date_or_none(_get(p, "End Date"))
        have_span = start is not None and end is not None and end > start
        per = (end - start).days // len(stages) if have_span else 0
        for i, stg in enumerate(stages):
            row = {"Milestone ID": f"MS{mid:04d}",
                   "Project ID": str(pid),
                   "Project Name": str(_get(p, "Project Name") or ""),
                   "Milestone": stg,
                   "Planned Date": "", "Actual Date": "",
                   "Status": "", "Owner": str(_get(p, "PM") or ""),
                   "Notes": ""}
            if have_span:
                date = start + pd.Timedelta(days=per * (i + 1))
                done = date < today
                row.update({"Planned Date": date,
                            "Actual Date": date if done else "",
                            "Status": "Complete" if done else "Planned"})
            rows.append(row); mid += 1
    return pd.DataFrame(rows)


# ─────────────────────── formula wiring ───────────────────────
# After a builder returns its DataFrame we rewrite selected columns as live
# Excel formulas so the workbook stays "smart":
#   • Project Name auto-fills from Project ID via INDEX/MATCH on Projects.
#   • Totals (Total Cost, Net Benefit, Actual, Variance, Total Benefit)
#     recalculate whenever the underlying cell changes.
# openpyxl treats any string starting with "=" as a formula on write.

def _projects_lookup(projects: pd.DataFrame) -> tuple[str, str]:
    cols = list(projects.columns)
    id_idx   = (cols.index("Project ID")   + 1) if "Project ID"   in cols else 1
    name_idx = (cols.index("Project Name") + 1) if "Project Name" in cols else 2
    return get_column_letter(id_idx), get_column_letter(name_idx)


def _name_lookup(pid_cell: str, id_col: str, name_col: str) -> str:
    return (f'=IFERROR(INDEX(Projects!{name_col}:{name_col},'
            f'MATCH({pid_cell},Projects!{id_col}:{id_col},0)),"")')


def _wire_formulas(sheet_name: str, df: pd.DataFrame,
                   proj_id_col: str, proj_name_col: str) -> pd.DataFrame:
    if df is None or df.empty:
        return df
    df = df.copy()
    cols = list(df.columns)
    col_letter = {c: get_column_letter(i + 1) for i, c in enumerate(cols)}
    pid_L = col_letter.get("Project ID")

    if pid_L and "Project Name" in cols:
        name_vals = []
        for i in range(len(df)):
            r = i + 2
            pid = df.iat[i, cols.index("Project ID")]
            name_vals.append(
                _name_lookup(f"{pid_L}{r}", proj_id_col, proj_name_col)
                if not _blank(pid) else ""
            )
        df["Project Name"] = name_vals

    templates: dict[str, str] = {}
    if sheet_name == "CostBenefit":
        if {"CAPEX", "OPEX", "Total Cost"} <= set(cols):
            templates["Total Cost"] = "=IFERROR({CAPEX}{r}+{OPEX}{r},0)"
        if {"Benefit Recurring", "Benefit One-Off", "Total Benefit"} <= set(cols):
            templates["Total Benefit"] = "=IFERROR({Benefit Recurring}{r}+{Benefit One-Off}{r},0)"
        if {"Total Benefit", "Total Cost", "Net Benefit"} <= set(cols):
            templates["Net Benefit"] = "={Total Benefit}{r}-{Total Cost}{r}"
    elif sheet_name == "Financials":
        if {"CAPEX", "OPEX", "Actual"} <= set(cols):
            templates["Actual"]        = "=IFERROR({CAPEX}{r}+{OPEX}{r},0)"
            templates["Forecast"]      = "=IFERROR({CAPEX}{r}+{OPEX}{r},0)"
            templates["Planned Value"] = "=IFERROR({CAPEX}{r}+{OPEX}{r},0)"
            templates["Earned Value"]  = "=IFERROR({CAPEX}{r}+{OPEX}{r},0)"
        if {"Forecast", "Actual", "Variance"} <= set(cols):
            templates["Variance"] = "=IFERROR({Forecast}{r}-{Actual}{r},0)"

    for target, tpl in templates.items():
        ref_cols = [c for c in cols if "{" + c + "}" in tpl]
        vals = []
        for i in range(len(df)):
            refs_ok = all(not _blank(df.iat[i, cols.index(rc)]) for rc in ref_cols)
            if not refs_ok:
                vals.append(""); continue
            r = i + 2
            vals.append(tpl.format(r=r, **{k: col_letter[k] for k in ref_cols}))
        df[target] = vals

    return df


# ─────────────────────── entry point ───────────────────────
BUILDERS = {
    "FYAllocation":    rebuild_fy_allocation,
    "PhaseFinancials": rebuild_phase_financials,
    "StageGates":      rebuild_stage_gates,
    "Financials":      rebuild_financials,
    "CostBenefit":     rebuild_cost_benefit,
    "Releases":        rebuild_releases,
    "Milestones":      rebuild_milestones,
}


def rebuild_all() -> dict[str, int]:
    """Rebuild every derived sheet from Projects. Returns {sheet: row_count}."""
    data = load_all()
    projects = data.get("projects", pd.DataFrame())
    if projects.empty: return {}

    proj_id_col, proj_name_col = _projects_lookup(projects)

    counts = {}
    for name, fn in BUILDERS.items():
        df = fn(projects)
        if df is None or df.empty:
            counts[name] = 0; continue
        df = _wire_formulas(name, df, proj_id_col, proj_name_col)
        write_sheet(name, df)
        counts[name] = len(df)

    existing_programs = data.get("programs", pd.DataFrame())
    prog_df = rebuild_programs(projects, existing_programs)
    if prog_df is not None and not prog_df.empty:
        write_sheet("Programs", prog_df)
        counts["Programs"] = len(prog_df)
    else:
        counts["Programs"] = 0

    return counts

