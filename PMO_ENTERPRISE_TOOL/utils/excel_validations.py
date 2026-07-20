"""Apply Excel data validations (dropdowns) to operational sheets, driven
entirely by the values in the workbook's `Config` sheet. Re-applied every
time a sheet is written so dropdowns always reflect the latest Config.

Validations use `error_style='warning'` so the user can still type new
values — the app remains tolerant to schema drift.
"""
from __future__ import annotations
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

# (sheet, column header) -> Config Category
TARGETS = {
    ("Projects", "Status"):             "Status",
    ("Projects", "RAG"):                "RAG",
    ("Projects", "Priority"):           "Priority",
    ("Projects", "Investment Type"):    "Investment Type",
    ("Projects", "Funding Type"):       "Funding Type",
    ("Projects", "Portfolio Category"): "Portfolio Category",
    ("Projects", "Governance Channel"): "Governance Channel",
    ("Projects", "Program"):            "Program",
    ("Projects", "Theme"):              "Theme",
    ("Projects", "Business Unit"):      "Business Unit",
    ("Projects", "Sponsor"):            "Sponsor",
    ("Projects", "Financial Year"):     "Financial Year",
    ("Governance", "Governance Channel"): "Governance Channel",
    ("Governance", "Gate Status"):      "Gate Status",
    ("Governance", "Gate Outcome"):     "Gate Outcome",
    ("StageGates", "Status"):           "Gate Status",
    ("StageGates", "Channel"):          "Governance Channel",
    ("Risks", "Probability"):           "Risk Probability",
    ("Risks", "Impact"):                "Risk Impact",
    ("Risks", "Status"):                "Risk Status",
    ("RAID", "Type"):                   "RAID Type",
    ("RAID", "Probability"):            "Risk Probability",
    ("RAID", "Impact"):                 "Risk Impact",
    ("RAID", "RAG"):                    "RAG",
    ("RAID", "Status"):                 "Action Status",
    ("Decisions", "Status"):            "Decision Status",
    ("Decisions", "Type"):              "Decision Type",
    ("Decisions", "Priority"):          "Risk Probability",
    ("Actions", "Status"):              "Action Status",
    ("Actions", "Priority"):            "Risk Probability",
    ("Benefits", "Status"):             "Benefit Status",
    ("Benefits", "Category"):           "Benefit Category",
    ("CostBenefit", "Benefit Type"):    "Benefit Type",
    ("CostBenefit", "Benefit Category"):"Benefit Category",
    ("Dependencies", "Dependency Type"):"Dependency Type",
    ("Dependencies", "Status"):         "Dependency Status",
    ("Dependencies", "Impact"):         "Risk Impact",
    ("Resources", "Skill"):             "Skill",
    ("Resources", "Role"):              "Resource Role",
    ("Pipeline", "Decision"):           "Pipeline Decision",
    ("Milestones", "Milestone"):        "Milestone",
    ("Milestones", "Status"):           "Action Status",
    ("PortfolioMovements", "From Category"): "Portfolio Category",
    ("PortfolioMovements", "To Category"):   "Portfolio Category",
}


def _lists_from_config(wb) -> dict[str, list[str]]:
    """Return {category: [active values...]} read from the Config sheet."""
    out: dict[str, list[str]] = {}
    if "Config" not in wb.sheetnames:
        return out
    ws = wb["Config"]
    headers = {str(c.value).strip(): i for i, c in enumerate(ws[1], 1) if c.value}
    cat_i = headers.get("Category"); val_i = headers.get("Value")
    act_i = headers.get("Active"); ord_i = headers.get("Display Order")
    if not cat_i or not val_i:
        return out
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        cat = r[cat_i - 1]; val = r[val_i - 1]
        if not cat or val is None:
            continue
        active = True if act_i is None else bool(r[act_i - 1])
        if not active:
            continue
        order = r[ord_i - 1] if ord_i else 0
        rows.append((str(cat).strip(), str(val), order if isinstance(order, (int, float)) else 0))
    rows.sort(key=lambda x: (x[0], x[2]))
    for cat, val, _ in rows:
        out.setdefault(cat, []).append(val)
    return out


def _col_letter(ws, header: str) -> str | None:
    for cell in ws[1]:
        if cell.value == header:
            return get_column_letter(cell.column)
    return None


def apply_validations_from_config(wb) -> None:
    """Apply (or refresh) dropdown validations on operational sheets."""
    lists = _lists_from_config(wb)
    if not lists:
        return
    for (sheet, header), category in TARGETS.items():
        if sheet not in wb.sheetnames:
            continue
        values = lists.get(category)
        if not values:
            continue
        ws = wb[sheet]
        col = _col_letter(ws, header)
        if not col:
            continue
        # Excel formula1 must be <=255 chars; collapse if needed
        formula = '"' + ",".join(v.replace('"', "'") for v in values) + '"'
        if len(formula) > 250:
            formula = '"' + ",".join(v.replace('"', "'") for v in values[:20]) + '"'
        dv = DataValidation(type="list", formula1=formula,
                            allow_blank=True, showDropDown=False,
                            errorStyle="warning",
                            error="Value not in Config; will still be saved.",
                            errorTitle="Not in Config list",
                            prompt=f"Pick from Config → {category}",
                            promptTitle=category)
        last_row = max(ws.max_row, 2)
        dv.add(f"{col}2:{col}{last_row + 500}")  # buffer for new rows
        ws.add_data_validation(dv)
