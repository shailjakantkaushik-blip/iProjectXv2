"""Edit any sheet of the active master Excel file directly in the app.

- Pick a sheet → grid is fully editable (add / delete / modify rows)
- Save writes back to the same .xlsx, preserving every other sheet
- All pages re-read the workbook automatically after save
"""
from __future__ import annotations
from pathlib import Path
from datetime import datetime
import shutil
import pandas as pd
import streamlit as st
from utils import auth as _auth; _auth.login_gate()
_auth.require_role(['admin'])
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

from config import get_master_file, DATA_DIR
from utils.excel_loader import load_all, refresh
from utils.theme_manager import apply_theme, render_sheet

apply_theme()


st.title("📝 Data Editor — edit Excel from the app")
st.caption("Changes save back to the source workbook. A timestamped backup is "
           "kept under `data/backups/` before every save.")

master = Path(get_master_file())
st.info(f"**Active file:** `{master}`", icon="📂")

if not master.exists():
    st.error(f"Master file not found at `{master}`. Set the active file from the **Data Source** page.")
    st.stop()

# Read fresh (bypass cache) so the editor always shows the on-disk truth.
try:
    xls = pd.ExcelFile(master, engine="openpyxl")
    sheet_names = list(xls.sheet_names)
except Exception as e:
    st.error(f"Could not open workbook: {e}")
    st.stop()

if not sheet_names:
    st.error("Workbook has no sheets.")
    st.stop()

sheet_name = st.selectbox("Sheet to edit", sheet_names, index=0)

try:
    original = pd.read_excel(xls, sheet_name=sheet_name)
except Exception as e:
    st.error(f"Could not read sheet `{sheet_name}`: {e}")
    st.stop()

# Normalize columns: stringify, dedupe, drop fully-unnamed trailing cols
new_cols: list[str] = []
seen: dict[str, int] = {}
for c in original.columns:
    name = str(c).strip()
    if not name or name.lower() == "nan" or name.startswith("Unnamed"):
        name = "Unnamed"
    if name in seen:
        seen[name] += 1
        name = f"{name}_{seen[name]}"
    else:
        seen[name] = 0
    new_cols.append(name)
original.columns = new_cols
original = original.reset_index(drop=True)

st.markdown(f"**Rows:** {len(original):,} &nbsp;·&nbsp; **Columns:** {len(original.columns):,}")

if original.empty or len(original.columns) == 0:
    st.warning("This sheet has no data. Edit the structure below to add columns/rows.")
    if len(original.columns) == 0:
        original = pd.DataFrame({"Column1": [""]})

def _cell_to_display(value) -> str:
    """Convert workbook values to stable editor text without turning data blank."""
    if pd.isna(value):
        return ""
    if isinstance(value, pd.Timestamp):
        return value.strftime("%Y-%m-%d") if value.time() == datetime.min.time() else value.isoformat(sep=" ")
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d") if value.time() == datetime.min.time() else value.isoformat(sep=" ")
    return str(value)


def _prepare_editor_df(df: pd.DataFrame) -> pd.DataFrame:
    """Use strings in the Streamlit grid to avoid blank canvas/render issues."""
    if df.empty and len(df.columns) == 0:
        return pd.DataFrame({"Column1": [""]})
    return df.astype("object").apply(lambda col: col.map(_cell_to_display))


def _restore_types(edited_df: pd.DataFrame, original_df: pd.DataFrame) -> pd.DataFrame:
    """Best-effort restore of numeric/date columns before writing back to Excel."""
    out = edited_df.copy()
    for col in out.columns:
        out[col] = out[col].replace("", pd.NA)
        if col not in original_df.columns:
            continue
        src = original_df[col]
        try:
            if pd.api.types.is_datetime64_any_dtype(src):
                out[col] = pd.to_datetime(out[col], errors="coerce")
            elif pd.api.types.is_integer_dtype(src):
                out[col] = pd.to_numeric(out[col], errors="coerce").astype("Int64")
            elif pd.api.types.is_float_dtype(src):
                out[col] = pd.to_numeric(out[col], errors="coerce")
        except Exception:
            pass
    return out


# Always show a non-canvas preview first, so sheet data is visible even when
# Streamlit's editable grid has a browser rendering issue.
st.markdown("**Sheet preview**")
render_sheet(original, max_rows=200)

editor_df = _prepare_editor_df(original)
st.markdown("**Editable sheet**")
edited = st.data_editor(
    editor_df,
    use_container_width=True,
    num_rows="dynamic",
    height=520,
    key=f"editor_{sheet_name}",
)


c1, c2, c3 = st.columns([1, 1, 4])
with c1:
    save = st.button("💾 Save to Excel", type="primary", use_container_width=True)
with c2:
    discard = st.button("↩️ Discard changes", use_container_width=True)

if discard:
    st.rerun()


def _backup(src: Path) -> Path:
    bdir = DATA_DIR / "backups"; bdir.mkdir(exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    dst = bdir / f"{src.stem}__{ts}{src.suffix}"
    shutil.copy2(src, dst)
    return dst


def _write_sheet(xlsx_path: Path, sheet: str, df: pd.DataFrame) -> None:
    """Replace one sheet in-place while preserving all other sheets.

    Uses openpyxl directly so formulas, formatting, and untouched sheets
    survive the round-trip.
    """
    wb = load_workbook(xlsx_path)
    if sheet in wb.sheetnames:
        # Preserve sheet position
        idx = wb.sheetnames.index(sheet)
        del wb[sheet]
        ws = wb.create_sheet(sheet, index=idx)
    else:
        ws = wb.create_sheet(sheet)
    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)
    wb.save(xlsx_path)


if save:
    try:
        from utils.progress import progress
        with progress(f"Saving {sheet_name} to Excel…", f"Saved {sheet_name}."):
            backup = _backup(master)
            restored = _restore_types(edited, original)
            to_save = restored.where(pd.notna(restored), "")
            _write_sheet(master, sheet_name, to_save)
            refresh()  # clear cache so every page sees the new data
        st.success(f"Saved **{sheet_name}** ({len(edited):,} rows). "
                   f"Backup: `{backup.name}`")
        st.balloons()
    except Exception as e:

        st.error(f"Save failed: {e}")

with st.expander("ℹ️ How saving works"):
    st.markdown(
        "- The selected sheet is **fully replaced** with the grid contents.\n"
        "- All other sheets in the workbook are left untouched.\n"
        "- A timestamped backup is written to `data/backups/` before every save.\n"
        "- After saving, the app's cache is cleared so every page reflects "
        "your changes on the next interaction.\n"
        "- If your workbook lives in a OneDrive / SharePoint-synced folder, "
        "the save propagates to SharePoint automatically via OneDrive sync."
    )
